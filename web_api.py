"""
Flask API — TG Channel Creator Web Panel
"""
import asyncio, json, os, re, zipfile, threading, time, shutil, uuid, secrets, hashlib
from datetime import datetime, timedelta
from pathlib import Path
from flask import Flask, jsonify, request, send_from_directory, Response
from flask_cors import CORS
from dotenv import load_dotenv

load_dotenv()

API_ID    = int(os.getenv("API_ID", "39759728"))
API_HASH  = os.getenv("API_HASH", "43cb9dc9f090c7f147d76eca7e6ebbbc")
BOT_TOKEN = os.getenv("BOT_TOKEN", "8192380712:AAFBEp99YqLpRKSvB_XIhyPH7ZXlprt9qTA")

SESSIONS_DIR    = "sessions"
CHANNELS_DIR    = "channels_data"
RESULTS_DIR     = "results"
NEURO_DIR       = "neuro_data"
SETTINGS_FILE   = "web_settings.json"
CHAN_AVATARS_DIR = "channel_avatars"
LICENSES_FILE    = "licenses.json"
WEB_SESSIONS     = "sessions.json"
USER_ACCOUNTS_FILE = "user_accounts.json"
ADMIN_MASTER     = os.getenv("ADMIN_MASTER", "2455228q")

for _d in [SESSIONS_DIR, CHANNELS_DIR, RESULTS_DIR, NEURO_DIR, CHAN_AVATARS_DIR]:
    Path(_d).mkdir(exist_ok=True)


def get_sessions_dir(username: str) -> str:
    """Возвращает путь к изолированной папке сессий пользователя."""
    d = Path(SESSIONS_DIR) / username
    d.mkdir(parents=True, exist_ok=True)
    return str(d)


def get_channels_dir(username: str) -> str:
    """Возвращает путь к изолированной папке данных каналов пользователя."""
    d = Path(CHANNELS_DIR) / username
    d.mkdir(parents=True, exist_ok=True)
    return str(d)

app = Flask(__name__, static_folder=".")
CORS(app)

# ── Логирование ────────────────────────────────────────────────────────────
import logging
import sys

log = logging.getLogger('werkzeug')
log.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stdout)
handler.setFormatter(logging.Formatter('[%(asctime)s] %(levelname)s: %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
log.addHandler(handler)

@app.after_request
def add_headers(resp):
    # Разрешаем clipboard API в браузере
    resp.headers["Permissions-Policy"] = "clipboard-read=*, clipboard-write=*"
    resp.headers["Cross-Origin-Opener-Policy"] = "same-origin"
    resp.headers["Cross-Origin-Embedder-Policy"] = "require-corp"
    return resp


# ── Лицензии / сессии ─────────────────────────────────────────────────────

def load_licenses() -> dict:
    try:
        return json.loads(Path(LICENSES_FILE).read_text(encoding="utf-8"))
    except Exception:
        return {}

def save_licenses(data: dict):
    Path(LICENSES_FILE).write_text(json.dumps(data, ensure_ascii=False, indent=2))

def load_web_sessions() -> dict:
    try:
        return json.loads(Path(WEB_SESSIONS).read_text(encoding="utf-8"))
    except Exception:
        return {}

def save_web_sessions(data: dict):
    Path(WEB_SESSIONS).write_text(json.dumps(data, ensure_ascii=False, indent=2))

def make_token() -> str:
    return "sess_" + secrets.token_hex(16)

def make_device_token() -> str:
    return "dt-" + secrets.token_hex(12)

def load_user_accounts() -> dict:
    try:
        return json.loads(Path(USER_ACCOUNTS_FILE).read_text(encoding="utf-8"))
    except Exception:
        return {}

def save_user_accounts(data: dict):
    Path(USER_ACCOUNTS_FILE).write_text(json.dumps(data, ensure_ascii=False, indent=2))

# Лимит попыток входа (in-memory, сбрасывается при рестарте)
_login_attempts: dict = {}  # {username: {"count": int, "blocked_until": float}}

def hash_password(pw: str) -> str:
    return hashlib.sha256(pw.encode("utf-8")).hexdigest()


# ── SMTP ─────────────────────────────────────────────────────────────────────

SMTP_DEFAULTS = {
    "smtp_host": "",
    "smtp_port": 587,
    "smtp_user": "",
    "smtp_password": "",
    "smtp_from_name": "TG Channel Creator",
    "smtp_tls": True,   # True = STARTTLS on port 587; False = SSL on port 465
}

def _smtp_cfg() -> dict:
    """Возвращает актуальные SMTP-настройки из settings."""
    return {k: settings.get(k, v) for k, v in SMTP_DEFAULTS.items()}

def email_template(subtitle: str, content_html: str) -> str:
    """Оборачивает контент в тёмный email-шаблон (table-based для совместимости)."""
    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#0c0e12;">
<table width="100%" cellpadding="0" cellspacing="0" border="0" bgcolor="#0c0e12">
  <tr><td align="center" style="padding:30px 16px;">
    <table width="480" cellpadding="0" cellspacing="0" border="0" bgcolor="#13161c"
           style="border-radius:12px;border:1px solid #1e2435;max-width:480px;">
      <tr><td style="padding:28px 28px 8px;">
        <div style="font-family:sans-serif;font-size:22px;font-weight:700;color:#4ade80;margin-bottom:6px;">TG Channel Creator</div>
        <div style="font-family:sans-serif;font-size:13px;color:#6b7280;">{subtitle}</div>
      </td></tr>
      <tr><td style="padding:16px 28px 28px;font-family:sans-serif;color:#e8eaf0;">
        {content_html}
      </td></tr>
    </table>
  </td></tr>
</table>
</body></html>"""

def send_email(to_addr: str, subject: str, body_html: str) -> tuple[bool, str]:
    """Отправляет письмо через SMTP. Возвращает (ok, error)."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    cfg = _smtp_cfg()
    host = cfg["smtp_host"].strip()
    port = int(cfg["smtp_port"] or 587)
    user = cfg["smtp_user"].strip()
    pwd  = cfg["smtp_password"]
    from_name = cfg["smtp_from_name"] or "TG Channel Creator"
    use_tls = cfg["smtp_tls"]
    if not host or not user:
        return False, "SMTP не настроен (укажите хост и логин в настройках)"
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = f"{from_name} <{user}>"
    msg["To"]      = to_addr
    msg.attach(MIMEText(body_html, "html", "utf-8"))
    try:
        if use_tls:
            with smtplib.SMTP(host, port, timeout=15) as s:
                s.ehlo()
                s.starttls()
                s.ehlo()
                if pwd:
                    s.login(user, pwd)
                s.sendmail(user, [to_addr], msg.as_string())
        else:
            with smtplib.SMTP_SSL(host, port, timeout=15) as s:
                s.ehlo()
                if pwd:
                    s.login(user, pwd)
                s.sendmail(user, [to_addr], msg.as_string())
        return True, ""
    except Exception as e:
        return False, str(e)[:200]


# ── Pending registrations & reset tokens ─────────────────────────────────────
# {email: {"username":..,"password_hash":..,"code":..,"expires":float,"attempts":int}}
_pending_regs: dict = {}
# {email: {"code":..,"expires":float,"attempts":int}}
_reset_tokens: dict = {}

def _clean_expired():
    now = time.time()
    for store in (_pending_regs, _reset_tokens):
        expired = [k for k, v in store.items() if v.get("expires", 0) < now]
        for k in expired:
            store.pop(k, None)

def _make_code() -> str:
    return str(secrets.randbelow(900000) + 100000)  # 6-значный код


# ── Проверка активного лицензионного ключа ────────────────────────────────

from functools import wraps

def _check_license_active(token: str) -> tuple[bool, str]:
    """
    Проверяет, что у пользователя с данным токеном есть активный (не истёкший) ключ.
    Возвращает (ok, error_message).
    """
    if not token:
        return False, "Требуется авторизация"
    sessions = load_web_sessions()
    if token not in sessions:
        return False, "Требуется авторизация"
    username = sessions[token].get("username")
    if not username:
        return False, "Сессия устарела"
    accounts = load_user_accounts()
    if username not in accounts:
        return False, "Пользователь не найден"
    lic_key = accounts[username].get("license_key")
    if not lic_key:
        return False, "Активный лицензионный ключ не привязан"
    licenses = load_licenses()
    if lic_key not in licenses:
        return False, "Ключ не найден в системе"
    exp = licenses[lic_key].get("expires_at")
    if exp:
        try:
            if datetime.fromisoformat(exp) < datetime.now():
                return False, "Срок действия лицензионного ключа истёк"
        except Exception:
            pass
    return True, ""


def require_license(f):
    """Декоратор: блокирует эндпоинт если у пользователя нет активного ключа.
    Устанавливает g.username для изоляции данных по пользователям."""
    @wraps(f)
    def decorated(*args, **kwargs):
        from flask import g
        token = request.headers.get("X-Token", "")
        print(f"[DEBUG] require_license: token={'<empty>' if not token else token[:20]}...", flush=True)
        sessions = load_web_sessions()
        print(f"[DEBUG] sessions.keys()={list(sessions.keys())}", flush=True)
        if token not in sessions:
            print(f"[DEBUG] token not found in sessions", flush=True)
            return jsonify({"ok": False, "error": "Нет доступа: Требуется авторизация"}), 403
        g.username = sessions[token].get("username", "")
        print(f"[DEBUG] g.username={g.username}", flush=True)
        ok, err = _check_license_active(token)
        print(f"[DEBUG] _check_license_active returned ok={ok}, err={err}", flush=True)
        if not ok:
            return jsonify({"ok": False, "error": f"Нет доступа: {err}"}), 403
        return f(*args, **kwargs)
    return decorated


state = {
    "running": False, "stop_flag": False,
    "progress": 0, "total": 0,
    "log": [], "results": [], "queue": [],
    "admin": "", "delay": 60, "description": "", "leave_creator": True,
    "broadcast_runs": 0, "broadcast_ok": 0, "broadcast_fail": 0,
    "broadcast_log": [], "broadcast_stop": False,
    "broadcast_running": False, "broadcast_sent": 0, "broadcast_total": 0,
    "check_running": False, "check_results": [],
    "neuro_log": [], "neuro_active": False,
}

settings = {
    "api_id": str(API_ID), "api_hash": API_HASH, "bot_token": BOT_TOKEN,
    "admin": "", "delay": 60, "msg_delay": 4, "max_channels": 7,
    "deepseek_api_key": "sk-3ddc1fd76de24ca7939fb99cd009ad0d",
    # SMTP
    "smtp_host": "smtp.gmail.com", "smtp_port": 587,
    "smtp_user": "goroinstag@gmail.com", "smtp_password": "viqa mzij hhib kluo",
    "smtp_from_name": "TG Channel Creator", "smtp_tls": True,
}

if Path(SETTINGS_FILE).exists():
    try:
        saved = json.loads(Path(SETTINGS_FILE).read_text(encoding="utf-8"))
        settings.update(saved)
        state["admin"] = saved.get("admin", "")
        state["delay"] = saved.get("delay", 60)
    except Exception:
        pass

# SMTP-настройки всегда принудительно (не перезаписываются из файла)
settings["smtp_host"]      = "smtp.gmail.com"
settings["smtp_port"]      = 587
settings["smtp_user"]      = "goroinstag@gmail.com"
settings["smtp_password"]  = "vdah xhya nejy kcgv"
settings["smtp_from_name"] = "TG Channel Creator"
settings["smtp_tls"]       = True


# ── Per-user state isolation ──────────────────────────────────────────────
# Каждый инструмент хранит state в словаре {username -> state_dict}
# Функции-геттеры создают state при первом обращении.

_u_state:     dict = {}
_u_acc_mgr:   dict = {}
_u_parser:    dict = {}
_u_inviter:   dict = {}
_u_pc:        dict = {}
_u_bcc:       dict = {}
_u_masslook:  dict = {}
_u_reactions: dict = {}
_u_reactions_last_ids: dict = {}   # username -> {ch -> last_id}
_u_reactions_done:     dict = {}   # username -> set of (ch, key)


def _new_main_state():
    return {
        "running": False, "stop_flag": False,
        "progress": 0, "total": 0,
        "log": [], "results": [], "queue": [],
        "admin": "", "delay": 60, "description": "", "leave_creator": True,
        "broadcast_runs": 0, "broadcast_ok": 0, "broadcast_fail": 0,
        "broadcast_log": [], "broadcast_stop": False,
        "broadcast_running": False, "broadcast_sent": 0, "broadcast_total": 0,
        "check_running": False, "check_results": [],
        "neuro_log": [], "neuro_active": False,
    }

def _new_acc_mgr_state():
    return {
        "spam_running": False, "spam_log": [], "spam_results": [],
        "spam_clean": 0, "spam_spam": 0, "spam_frozen": 0,
        "clean_running": False, "clean_stop": False,
        "clean_log": [], "clean_progress": 0, "clean_total": 0,
    }

def _new_parser_state():
    return {
        "running": False, "stop_flag": False,
        "progress": 0, "total": 0,
        "found": 0, "filtered": 0,
        "log": [], "results": [],
    }

def _new_inviter_state():
    return {
        "running": False, "stop_flag": False,
        "progress": 0, "total": 0,
        "ok": 0, "fail": 0, "swaps": 0,
        "log": [], "chat_stats": {},
        "cur_acc": "", "cur_chat": "",
        "prepare_running": False, "prepare_done": False,
        "prepare_phase": "",
        "prepare_progress": 0, "prepare_total": 0,
        "prepare_detail": "",
    }

def _new_pc_state():
    return {
        "running": False, "stop_flag": False,
        "progress": 0, "total": 0,
        "found": 0, "notfound": 0, "errors": 0,
        "log": [], "results": [], "cur": "",
    }

def _new_bcc_state():
    return {
        "running": False, "stop_flag": False,
        "progress": 0, "total": 0,
        "sent": 0, "fail": 0, "skip": 0,
        "log": [], "cur": "",
    }

def _new_masslook_state():
    return {
        "running": False, "stop_flag": False,
        "progress": 0, "total": 0,
        "viewed": 0, "liked": 0, "users_done": 0, "errors": 0,
        "log": [], "cur_acc": "", "cur_user": "",
    }

def _new_reactions_state():
    return {"active": False, "stop_flag": False, "sent": 0, "errors": 0, "joined": 0, "log": []}


def _ust(u):
    if u not in _u_state:     _u_state[u]    = _new_main_state()
    return _u_state[u]

def _uam(u):
    if u not in _u_acc_mgr:   _u_acc_mgr[u]  = _new_acc_mgr_state()
    return _u_acc_mgr[u]

def _uprs(u):
    if u not in _u_parser:    _u_parser[u]   = _new_parser_state()
    return _u_parser[u]

def _uinv(u):
    if u not in _u_inviter:   _u_inviter[u]  = _new_inviter_state()
    return _u_inviter[u]

def _upc(u):
    if u not in _u_pc:        _u_pc[u]       = _new_pc_state()
    return _u_pc[u]

def _ubcc(u):
    if u not in _u_bcc:       _u_bcc[u]      = _new_bcc_state()
    return _u_bcc[u]

def _uml(u):
    if u not in _u_masslook:  _u_masslook[u] = _new_masslook_state()
    return _u_masslook[u]

def _ureact(u):
    if u not in _u_reactions:
        _u_reactions[u]           = _new_reactions_state()
        _u_reactions_last_ids[u]  = {}
        _u_reactions_done[u]      = set()
    return _u_reactions[u]


def _ulog(lst, level, msg, prefix=""):
    entry = {"ts": datetime.now().strftime("%H:%M:%S"), "level": level, "msg": msg}
    lst.append(entry)
    if len(lst) > 600:
        del lst[:-400]
    if prefix:
        print(f"[{prefix}/{level.upper()}] {msg}")
    else:
        print(f"[{level.upper()}] {msg}")


def log(level, msg):
    entry = {"ts": datetime.now().strftime("%H:%M:%S"), "level": level, "msg": msg}
    state["log"].append(entry)
    if len(state["log"]) > 600:
        state["log"] = state["log"][-400:]
    print(f"[{level.upper()}] {msg}")


# ── Авторизация через веб ─────────────────────────────────────────────────

auth_sessions = {}  # phone -> {"client": ..., "phone_code_hash": ..., "loop": ..., "thread": ...}


def _run_auth_coroutine(phone, coro):
    """Запускает корутину в loop авторизационной сессии."""
    sess = auth_sessions.get(phone)
    if sess and sess.get("loop"):
        loop = sess["loop"]
        future = asyncio.run_coroutine_threadsafe(coro, loop)
        return future.result(timeout=30)
    else:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            return loop.run_until_complete(coro)
        finally:
            loop.close()


@app.route("/api/auth/send_code", methods=["POST"])
@require_license
def api_auth_send_code():
    from flask import g
    data  = (request.get_json(force=True, silent=True) or {})
    phone = data.get("phone", "").strip().lstrip("+")
    if not phone:
        return jsonify({"ok": False, "error": "Укажите номер телефона"})
    _owner = g.username

    # Убиваем старую сессию
    old = auth_sessions.pop(phone, None)
    if old:
        try:
            old_loop = old.get("loop")
            old_client = old.get("client")
            if old_loop and old_client:
                asyncio.run_coroutine_threadsafe(old_client.disconnect(), old_loop)
        except Exception:
            pass

    result_holder = {}
    ready = threading.Event()

    def _thread():
        from telethon import TelegramClient, errors
        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

        async def _run():
            session_file = str(Path(get_sessions_dir(_owner)) / phone)
            client = TelegramClient(session_file, api_id, api_hash)
            try:
                await client.connect()
                result = await client.send_code_request(phone)
                auth_sessions[phone] = {
                    "client": client,
                    "phone_code_hash": result.phone_code_hash,
                    "loop": loop,
                    "owner": _owner,
                }
                log("info", f"Код отправлен на +{phone}")
                result_holder["res"] = {"ok": True}
            except errors.FloodWaitError as fe:
                await client.disconnect()
                result_holder["res"] = {"ok": False, "error": f"FloodWait {fe.seconds}s — попробуйте позже"}
            except errors.PhoneNumberInvalidError:
                await client.disconnect()
                result_holder["res"] = {"ok": False, "error": "Неверный номер телефона"}
            except Exception as ex:
                try: await client.disconnect()
                except: pass
                result_holder["res"] = {"ok": False, "error": str(ex)[:120]}
            finally:
                ready.set()

        loop.run_until_complete(_run())
        # Если сессия создана — держим loop живым для следующих вызовов
        if phone in auth_sessions:
            loop.run_forever()
        loop.close()

    t = threading.Thread(target=_thread, daemon=True)
    t.start()
    ready.wait(timeout=30)
    return jsonify(result_holder.get("res", {"ok": False, "error": "Таймаут"}))


@app.route("/api/auth/sign_in", methods=["POST"])
@require_license
def api_auth_sign_in():
    data  = (request.get_json(force=True, silent=True) or {})
    phone = data.get("phone", "").strip().lstrip("+")
    code  = data.get("code", "").strip()
    if not phone or not code:
        return jsonify({"ok": False, "error": "Нет телефона или кода"})

    sess = auth_sessions.get(phone)
    if not sess:
        return jsonify({"ok": False, "error": "Сначала запросите код — сессия не найдена"})

    client = sess["client"]
    loop   = sess["loop"]
    phone_code_hash = sess["phone_code_hash"]
    _owner = sess.get("owner", "")

    async def _sign():
        from telethon import errors
        try:
            me = await client.sign_in(phone=phone, code=code, phone_code_hash=phone_code_hash)
            await client.disconnect()
            auth_sessions.pop(phone, None)
            loop.call_soon_threadsafe(loop.stop)
            _save_cache(phone, me, _owner)
            log("success", f"+{phone} авторизован: {me.first_name}")
            return {"ok": True,
                    "name": (me.first_name or "") + (" " + (me.last_name or "") if me.last_name else ""),
                    "username": me.username or ""}
        except errors.SessionPasswordNeededError:
            return {"ok": False, "need_2fa": True}
        except errors.PhoneCodeInvalidError:
            return {"ok": False, "error": "Неверный код"}
        except errors.PhoneCodeExpiredError:
            return {"ok": False, "error": "Код истёк — запросите новый"}
        except Exception as ex:
            return {"ok": False, "error": str(ex)[:120]}

    try:
        future = asyncio.run_coroutine_threadsafe(_sign(), loop)
        result = future.result(timeout=30)
    except Exception as ex:
        result = {"ok": False, "error": str(ex)[:120]}
    return jsonify(result)


@app.route("/api/auth/sign_in_2fa", methods=["POST"])
@require_license
def api_auth_sign_in_2fa():
    data     = (request.get_json(force=True, silent=True) or {})
    phone    = data.get("phone", "").strip().lstrip("+")
    password = data.get("password", "")
    if not phone or not password:
        return jsonify({"ok": False, "error": "Нет телефона или пароля"})

    sess = auth_sessions.get(phone)
    if not sess:
        return jsonify({"ok": False, "error": "Сессия не найдена — начните заново"})

    client = sess["client"]
    loop   = sess["loop"]
    _owner = sess.get("owner", "")

    async def _sign_2fa():
        from telethon import errors
        try:
            me = await client.sign_in(password=password)
            await client.disconnect()
            auth_sessions.pop(phone, None)
            loop.call_soon_threadsafe(loop.stop)
            _save_cache(phone, me, _owner)
            log("success", f"+{phone} авторизован (2FA): {me.first_name}")
            return {"ok": True,
                    "name": (me.first_name or "") + (" " + (me.last_name or "") if me.last_name else ""),
                    "username": me.username or ""}
        except errors.PasswordHashInvalidError:
            return {"ok": False, "error": "Неверный пароль 2FA"}
        except Exception as ex:
            return {"ok": False, "error": str(ex)[:120]}

    try:
        future = asyncio.run_coroutine_threadsafe(_sign_2fa(), loop)
        result = future.result(timeout=30)
    except Exception as ex:
        result = {"ok": False, "error": str(ex)[:120]}
    return jsonify(result)


def _save_cache(phone, me, owner=""):
    try:
        cache = {
            "status": "valid",
            "name": (me.first_name or "") + (" " + me.last_name if me.last_name else ""),
            "username": me.username or ""
        }
        (Path(get_sessions_dir(owner)) / f"{phone}.cache").write_text(
            json.dumps(cache, ensure_ascii=False))
    except Exception:
        pass


# ── Аккаунты ──────────────────────────────────────────────────────────────

def get_accounts(owner: str = ""):
    """Возвращает список аккаунтов пользователя owner (изолированно)."""
    sdir = Path(get_sessions_dir(owner)) if owner else Path(SESSIONS_DIR)
    cdir = Path(get_channels_dir(owner)) if owner else Path(CHANNELS_DIR)
    accounts = []
    for sf in sorted(sdir.glob("*.session")):
        phone = sf.stem
        channels = []
        meta = cdir / f"{phone}.json"
        if meta.exists():
            try:
                channels = json.loads(meta.read_text(encoding="utf-8")).get("channels", [])
            except Exception:
                pass
        status, name, uname, tag, premium = "unknown", "", "", "", False
        cache_f = sdir / f"{phone}.cache"
        if cache_f.exists():
            try:
                c = json.loads(cache_f.read_text(encoding="utf-8"))
                status  = c.get("status", "unknown")
                name    = c.get("name", "")
                uname   = c.get("username", "")
                tag     = c.get("tag", "")
                premium = c.get("premium", False)
            except Exception:
                pass
        else:
            old_f = sdir / f"{phone}.status"
            if old_f.exists():
                try: status = old_f.read_text().strip()
                except Exception: pass
        accounts.append({"phone": phone, "channels": len(channels),
                         "status": status, "name": name, "username": uname,
                         "tag": tag, "premium": premium})
    return accounts


@app.route("/api/accounts")
@require_license
def api_accounts():
    from flask import g
    return jsonify({"accounts": get_accounts(g.username)})


@app.route("/api/accounts/delete", methods=["POST"])
@require_license
def api_delete_account():
    from flask import g
    sdir = Path(get_sessions_dir(g.username))
    cdir = Path(get_channels_dir(g.username))
    phone = ((request.get_json(force=True, silent=True) or {})).get("phone", "")
    for pat in [f"{phone}.session", f"{phone}.session.journal"]:
        p = sdir / pat
        if p.exists(): p.unlink()
    for pat in [cdir / f"{phone}.json",
                sdir / f"{phone}.status",
                sdir / f"{phone}.cache"]:
        if pat.exists(): pat.unlink()
    log("warn", f"Аккаунт {phone} удалён")
    return jsonify({"ok": True})


@app.route("/api/accounts/reconnect", methods=["POST"])
@require_license
def api_reconnect():
    from flask import g
    accs = get_accounts(g.username)
    log("info", f"Переподключение {len(accs)} аккаунтов...")
    log("success", f"Готово. Аккаунтов: {len(accs)}")
    return jsonify({"ok": True, "count": len(accs)})


# ── ZIP загрузка (ИСПРАВЛЕНО) ──────────────────────────────────────────────

@app.route("/api/accounts/upload_zip", methods=["POST"])
@require_license
def api_upload_zip():
    from flask import g
    _owner = g.username
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Файл не найден в запросе"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "error": "Пустое имя файла"}), 400

    tmp = Path("tmp_upload.zip")
    try:
        f.save(str(tmp))
    except Exception as ex:
        return jsonify({"ok": False, "error": f"Не удалось сохранить: {ex}"}), 500

    extracted = []
    errors = []
    sdir = Path(get_sessions_dir(_owner))

    try:
        with zipfile.ZipFile(str(tmp), "r") as z:
            all_files = z.namelist()
            # Ищем .session файлы на любой глубине, пропускаем директории
            session_files = [
                n for n in all_files
                if n.endswith(".session") and not n.endswith("/")
                and not z.getinfo(n).is_dir()
            ]

            if not session_files:
                return jsonify({"ok": False, "error": f"В архиве нет .session файлов. Найдено: {all_files[:5]}"}), 400

            for name in session_files:
                try:
                    # Получаем только имя файла без папок (поддерживаем вложенную структуру)
                    basename = Path(name).name
                    phone = Path(basename).stem

                    # Читаем содержимое и сохраняем в папку пользователя
                    data = z.read(name)
                    if not data:
                        errors.append(f"{basename}: пустой файл")
                        continue
                    dest = sdir / basename
                    dest.write_bytes(data)

                    extracted.append(phone)
                    log("success", f"Сессия извлечена: {phone}")
                except Exception as ex:
                    errors.append(f"{basename}: {type(ex).__name__}: {ex}")
                    log("error", f"Ошибка извлечения {name}: {type(ex).__name__}: {ex}")

    except zipfile.BadZipFile:
        return jsonify({"ok": False, "error": "Файл не является ZIP архивом"}), 400
    except Exception as ex:
        log("error", f"ZIP ошибка: {ex}")
        return jsonify({"ok": False, "error": str(ex)}), 500
    finally:
        if tmp.exists():
            tmp.unlink()

    if not extracted:
        return jsonify({"ok": False, "error": f"Не удалось извлечь ни одной сессии. Ошибки: {errors}"}), 400

    return jsonify({"ok": True, "extracted": extracted, "count": len(extracted), "errors": errors})


# ── Чекер валидности аккаунтов ─────────────────────────────────────────────

def check_accounts_task(st):
    async def _check():
        try:
            from telethon import TelegramClient, errors
        except ImportError:
            log("error", "Telethon не установлен — pip install telethon")
            st["check_running"] = False
            return

        _uname = st.get("_username", "")
        accounts = get_accounts(_uname)
        if not accounts:
            log("warn", "Нет аккаунтов для проверки")
            st["check_running"] = False
            return

        log("info", f"Чекер запущен. Проверяю {len(accounts)} аккаунтов...")
        st["check_results"] = []
        api_id  = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)

        sdir_check = get_sessions_dir(_uname)
        for acc in accounts:
            phone = acc["phone"]
            session_file = str(Path(sdir_check) / phone)
            result = {"phone": phone, "status": "unknown", "name": "", "username": "", "error": ""}
            try:
                client = TelegramClient(session_file, api_id, api_hash)
                await client.connect()
                if await client.is_user_authorized():
                    me = await client.get_me()
                    result["status"] = "valid"
                    result["name"] = (me.first_name or "") + (" " + me.last_name if me.last_name else "")
                    result["username"] = me.username or ""
                    log("success", f"✓ {phone} — {result['name']} (@{result['username']})")
                else:
                    result["status"] = "invalid"
                    result["error"] = "Сессия не авторизована"
                    log("warn", f"✗ {phone} — не авторизован")
                await client.disconnect()
            except errors.AuthKeyUnregisteredError:
                result["status"] = "dead"
                result["error"] = "Ключ авторизации отозван (аккаунт забанен/удалён)"
                log("error", f"✗ {phone} — ключ отозван")
            except errors.UserDeactivatedError:
                result["status"] = "dead"
                result["error"] = "Аккаунт удалён"
                log("error", f"✗ {phone} — аккаунт удалён")
            except errors.FloodWaitError as fe:
                result["status"] = "flood"
                result["error"] = f"FloodWait {fe.seconds}s"
                log("warn", f"⏳ {phone} — FloodWait {fe.seconds}s")
                await asyncio.sleep(min(fe.seconds, 30))
            except FileNotFoundError:
                result["status"] = "missing"
                result["error"] = "Файл сессии не найден"
                log("error", f"✗ {phone} — файл не найден")
            except Exception as ex:
                result["status"] = "error"
                result["error"] = str(ex)[:120]
                log("error", f"✗ {phone} — {ex}")

            # Кэшируем результат
            try:
                cache_data = {"status": result["status"], "name": result["name"], "username": result["username"]}
                (Path(sdir_check) / f"{phone}.cache").write_text(
                    json.dumps(cache_data, ensure_ascii=False))
            except Exception:
                pass

            st["check_results"].append(result)
            await asyncio.sleep(1.5)

        valid = sum(1 for r in st["check_results"] if r["status"] == "valid")
        dead  = sum(1 for r in st["check_results"] if r["status"] in ("invalid","dead"))
        log("success", f"Чекер завершён. Валидных: {valid}, мёртвых: {dead}")
        st["check_running"] = False

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(_check())
    loop.close()


@app.route("/api/accounts/check", methods=["POST"])
@require_license
def api_check_accounts():
    from flask import g
    st = _ust(g.username)
    if st["check_running"]:
        return jsonify({"ok": False, "error": "Проверка уже запущена"}), 400
    st["check_running"] = True
    st["check_results"] = []
    st["_username"] = g.username
    threading.Thread(target=check_accounts_task, args=(st,), daemon=True).start()
    _ulog(st["log"], "info", "Запуск чекера аккаунтов...")
    return jsonify({"ok": True})


@app.route("/api/accounts/check/status")
@require_license
def api_check_status():
    from flask import g
    st = _ust(g.username)
    return jsonify({
        "running": st["check_running"],
        "results": st["check_results"],
        "total": len(get_accounts(g.username)),
        "done": len(st["check_results"]),
    })


# ── Менеджер аккаунтов ────────────────────────────────────────────────────

acc_mgr_state = {
    "spam_running": False, "spam_log": [], "spam_results": [],
    "spam_clean": 0, "spam_spam": 0, "spam_frozen": 0,
    "clean_running": False, "clean_stop": False,
    "clean_log": [], "clean_progress": 0, "clean_total": 0,
}


def acc_mgr_log(st, key, level, msg):
    _ulog(st[key], level, msg, "ACC_MGR")


# ── Спамблок ──────────────────────────────────────────────────────────────

def spam_check_task(phones, st):
    try:
        import asyncio
        from telethon import TelegramClient, errors

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)
        _uname = st.get("_username", "")
        sdir = get_sessions_dir(_uname)
        accounts = get_accounts(_uname)
        if phones:
            accounts = [a for a in accounts if a["phone"] in phones]

        if not accounts:
            acc_mgr_log(st, "spam_log", "warn", "Нет аккаунтов для проверки")
            st["spam_running"] = False
            return

        st["spam_results"] = []
        st["spam_clean"]   = 0
        st["spam_spam"]    = 0
        st["spam_frozen"]  = 0

        acc_mgr_log(st, "spam_log", "info", f"Начинаю проверку {len(accounts)} аккаунтов...")

        async def check_one(phone):
            sf = str(Path(sdir) / phone)
            result = {"phone": phone, "status": "unknown", "status_text": "Неизвестно"}
            client = None
            try:
                client = TelegramClient(sf, api_id, api_hash)
                await client.connect()

                if not await client.is_user_authorized():
                    result = {"phone": phone, "status": "dead", "status_text": "Сессия мертва"}
                    acc_mgr_log(st, "spam_log", "error", f"[{phone[-4:]}] Сессия мертва")
                    return result

                acc_mgr_log(st, "spam_log", "info", f"[{phone[-4:]}] Открываю @SpamBot...")

                # Получаем SpamBot entity
                spambot = await client.get_entity("SpamBot")

                # Отправляем /start и ждём ответ
                await client.send_message(spambot, "/start")
                await asyncio.sleep(6)

                # Читаем последние сообщения от бота (только входящие)
                msgs = await client.get_messages(spambot, limit=3)
                bot_msgs = [m for m in msgs if not m.out and m.text]

                if not bot_msgs:
                    # Попробуем ещё подождать
                    await asyncio.sleep(4)
                    msgs = await client.get_messages(spambot, limit=3)
                    bot_msgs = [m for m in msgs if not m.out and m.text]

                text = " ".join(m.text for m in bot_msgs).lower()
                acc_mgr_log(st, "spam_log", "info", f"[{phone[-4:]}] Ответ: {text[:120]}")

                if not text:
                    result = {"phone": phone, "status": "unknown", "status_text": "Нет ответа от @SpamBot"}
                    acc_mgr_log(st, "spam_log", "warn", f"[{phone[-4:]}] Нет ответа от бота")
                elif any(k in text for k in [
                    "no limits", "нет ограничений", "good standing",
                    "hasn't been limited", "not been limited",
                    "не имеет ограничений", "ограничений нет", "free to use",
                    "your account is clear"
                ]):
                    result = {"phone": phone, "status": "clean", "status_text": "✓ Чистый"}
                    st["spam_clean"] += 1
                    acc_mgr_log(st, "spam_log", "success", f"[{phone[-4:]}] Чистый ✓")

                elif any(k in text for k in [
                    "spam", "limited", "ограничен", "restrict",
                    "report", "жалоб", "пожаловались", "reported",
                    "too many requests"
                ]):
                    result = {"phone": phone, "status": "spam", "status_text": "🚫 Спамблок"}
                    st["spam_spam"] += 1
                    acc_mgr_log(st, "spam_log", "warn", f"[{phone[-4:]}] СПАМБЛОК 🚫")

                elif any(k in text for k in [
                    "deactivated", "заморожен", "frozen", "deleted", "banned"
                ]):
                    result = {"phone": phone, "status": "frozen", "status_text": "❄ Заморожен"}
                    st["spam_frozen"] += 1
                    acc_mgr_log(st, "spam_log", "error", f"[{phone[-4:]}] Заморожен ❄")

                else:
                    raw = text[:100] if text else "(нет ответа)"
                    result = {"phone": phone, "status": "unknown", "status_text": f"? {raw}"}
                    acc_mgr_log(st, "spam_log", "warn", f"[{phone[-4:]}] Неизвестный ответ: {raw}")

            except errors.UserDeactivatedError:
                result = {"phone": phone, "status": "frozen", "status_text": "❄ Заморожен"}
                st["spam_frozen"] += 1
                acc_mgr_log(st, "spam_log", "error", f"[{phone[-4:]}] Заморожен (UserDeactivated)")
            except errors.FloodWaitError as fe:
                result = {"phone": phone, "status": "error", "status_text": f"FloodWait {fe.seconds}s"}
                acc_mgr_log(st, "spam_log", "warn", f"[{phone[-4:]}] FloodWait {fe.seconds}s — жду...")
                await asyncio.sleep(min(fe.seconds, 30))
            except Exception as ex:
                result = {"phone": phone, "status": "error", "status_text": str(ex)[:60]}
                acc_mgr_log(st, "spam_log", "error", f"[{phone[-4:]}] Ошибка: {str(ex)[:120]}")
            finally:
                if client:
                    try: await client.disconnect()
                    except: pass
            return result

        async def run_all():
            for acc in accounts:
                result = await check_one(acc["phone"])
                st["spam_results"].append(result)
                await asyncio.sleep(3)

            st["spam_running"] = False
            acc_mgr_log(st, "spam_log", "success",
                f"=== Готово. Чистых: {st['spam_clean']} | "
                f"Спам: {st['spam_spam']} | "
                f"Заморожено: {st['spam_frozen']} ==="
            )

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            loop.run_until_complete(run_all())
        finally:
            loop.close()

    except ImportError:
        acc_mgr_log(st, "spam_log", "error", "Telethon не установлен: pip install telethon")
        st["spam_running"] = False
    except Exception as ex:
        acc_mgr_log(st, "spam_log", "error", f"Критическая ошибка: {str(ex)}")
        st["spam_running"] = False


@app.route("/api/accounts/spam_check", methods=["POST"])
@require_license
def api_spam_check():
    from flask import g
    st = _uam(g.username)

    # Если уже запущена — принудительно сбрасываем флаг и запускаем заново
    if st["spam_running"]:
        st["spam_running"] = False
        import time; time.sleep(0.3)
    acc_mgr_state.update({"spam_running": True, "spam_log": [], "spam_results": [],
                           "spam_clean": 0, "spam_spam": 0, "spam_frozen": 0})
    try:
        data = request.get_json(force=True, silent=True) or {}
    except Exception:
        data = {}
    phones = data.get("phones", [])
    st["_username"] = g.username
    threading.Thread(target=spam_check_task, args=(phones, st), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/accounts/spam_reset", methods=["POST"])
@require_license
def api_spam_reset():
    from flask import g
    st = _uam(g.username)

    """Принудительный сброс флага зависшей проверки."""
    st["spam_running"] = False
    return jsonify({"ok": True})


@app.route("/api/accounts/spam_status")
@require_license
def api_spam_status():
    from flask import g
    st = _uam(g.username)

    return jsonify({
        "running": st["spam_running"],
        "log":     st["spam_log"],
        "results": st["spam_results"],
        "clean":   st["spam_clean"],
        "spam":    st["spam_spam"],
        "frozen":  st["spam_frozen"],
    })


@app.route("/api/accounts/spam_unban", methods=["POST"])
@require_license
def api_spam_unban():
    from flask import g
    st = _uam(g.username)

    async def _run():
        try:
            from telethon import TelegramClient, errors
            from telethon.tl.functions.messages import GetBotCallbackAnswerRequest
        except ImportError:
            return {"ok": False, "error": "Telethon не установлен"}

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)
        # Только те у кого спамблок
        targets = [r for r in st["spam_results"] if r["status"] == "spam"]
        if not targets:
            return {"ok": False, "error": "Нет аккаунтов со спамблоком — сначала запустите проверку"}

        count = 0
        for r in targets:
            phone = r["phone"]
            sf = str(Path(get_sessions_dir(st.get("_username",""))) / phone)
            client = None
            try:
                client = TelegramClient(sf, api_id, api_hash)
                await client.connect()
                if not await client.is_user_authorized(): continue
                spambot = await client.get_entity("SpamBot")
                await client.send_message(spambot, "/start")
                await asyncio.sleep(2)
                msgs = await client.get_messages(spambot, limit=1)
                if msgs and msgs[0].reply_markup:
                    for row in msgs[0].reply_markup.rows:
                        for btn in row.buttons:
                            if hasattr(btn, 'data'):
                                try:
                                    await client(GetBotCallbackAnswerRequest(
                                        peer=spambot, msg_id=msgs[0].id, data=btn.data))
                                    count += 1
                                except Exception:
                                    pass
            except Exception as ex:
                log("warn", f"[{phone[-4:]}] Unban ошибка: {str(ex)[:60]}")
            finally:
                if client:
                    try: await client.disconnect()
                    except: pass
            await asyncio.sleep(3)
        return {"ok": True, "count": count}

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    result = loop.run_until_complete(_run())
    loop.close()
    return jsonify(result)


# ── Редактирование профиля ────────────────────────────────────────────────

@app.route("/api/accounts/edit_profile", methods=["POST"])
@require_license
def api_edit_profile():
    from flask import g
    data       = (request.get_json(force=True, silent=True) or {})
    phone      = data.get("phone", "").strip()
    first_name = data.get("first_name", "")
    last_name  = data.get("last_name", "")
    bio        = data.get("bio", "")
    username   = data.get("username", "").lstrip("@").strip()
    _owner     = g.username

    async def _run():
        try:
            from telethon import TelegramClient, errors
            from telethon.tl.functions.account import UpdateProfileRequest, UpdateUsernameRequest
        except ImportError:
            return {"ok": False, "error": "Telethon не установлен"}

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)
        sf = str(Path(get_sessions_dir(_owner)) / phone)
        client = None
        try:
            client = TelegramClient(sf, api_id, api_hash)
            await client.connect()
            if not await client.is_user_authorized():
                return {"ok": False, "error": "Сессия не авторизована"}

            if first_name or last_name or bio:
                await client(UpdateProfileRequest(
                    first_name=first_name,
                    last_name=last_name,
                    about=bio
                ))

            if username:
                await client(UpdateUsernameRequest(username=username))

            # Обновляем кэш
            me = await client.get_me()
            cache = {
                "status": "valid",
                "name": (me.first_name or "") + (" " + me.last_name if me.last_name else ""),
                "username": me.username or ""
            }
            (Path(get_sessions_dir(_owner)) / f"{phone}.cache").write_text(
                json.dumps(cache, ensure_ascii=False))
            log("success", f"[{phone[-4:]}] Профиль обновлён")
            return {"ok": True}
        except errors.UsernameOccupiedError:
            return {"ok": False, "error": "Username уже занят"}
        except errors.UsernameInvalidError:
            return {"ok": False, "error": "Недопустимый username"}
        except Exception as ex:
            return {"ok": False, "error": str(ex)[:120]}
        finally:
            if client:
                try: await client.disconnect()
                except: pass

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    result = loop.run_until_complete(_run())
    loop.close()
    return jsonify(result)


@app.route("/api/accounts/upload_avatar", methods=["POST"])
@require_license
def api_upload_avatar():
    from flask import g
    phone = request.form.get("phone", "").strip()
    _owner = g.username
    if "photo" not in request.files or not phone:
        return jsonify({"ok": False, "error": "Нет файла или телефона"})

    photo_file = request.files["photo"]
    tmp_path   = Path("tmp_avatar_" + phone + Path(photo_file.filename).suffix)
    photo_file.save(str(tmp_path))

    async def _run():
        try:
            from telethon import TelegramClient
            from telethon.tl.functions.photos import UploadProfilePhotoRequest
        except ImportError:
            return {"ok": False, "error": "Telethon не установлен"}

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)
        sf = str(Path(get_sessions_dir(_owner)) / phone)
        client = None
        try:
            client = TelegramClient(sf, api_id, api_hash)
            await client.connect()
            if not await client.is_user_authorized():
                return {"ok": False, "error": "Сессия не авторизована"}
            uploaded = await client.upload_file(str(tmp_path))
            await client(UploadProfilePhotoRequest(file=uploaded))
            log("success", f"[{phone[-4:]}] Аватарка загружена")
            return {"ok": True}
        except Exception as ex:
            return {"ok": False, "error": str(ex)[:120]}
        finally:
            if client:
                try: await client.disconnect()
                except: pass
            try: tmp_path.unlink()
            except: pass

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    result = loop.run_until_complete(_run())
    loop.close()
    return jsonify(result)


# ── Очистка чатов ─────────────────────────────────────────────────────────

def clean_task(phones, groups, channels, bots, delay, st):
    async def _run():
        try:
            from telethon import TelegramClient, errors
            from telethon.tl.functions.channels import LeaveChannelRequest
            from telethon.tl.functions.messages import DeleteChatUserRequest
            from telethon.tl.types import Channel, Chat
        except ImportError:
            acc_mgr_log(st, "clean_log", "error", "Telethon не установлен")
            st["clean_running"] = False
            return

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)
        _uname = st.get("_username", "")
        sdir = get_sessions_dir(_uname)
        accounts = get_accounts(_uname)
        if phones:
            accounts = [a for a in accounts if a["phone"] in phones]

        acc_mgr_log(st, "clean_log", "info", f"Очистка {len(accounts)} аккаунтов...")

        for acc in accounts:
            if st["clean_stop"]: break
            phone  = acc["phone"]
            sf     = str(Path(sdir) / phone)
            client = None
            try:
                client = TelegramClient(sf, api_id, api_hash)
                await client.connect()
                if not await client.is_user_authorized():
                    acc_mgr_log(st, "clean_log", "warn", f"[{phone[-4:]}] Не авторизован, пропуск")
                    continue

                dialogs = await client.get_dialogs()
                total = len(dialogs)
                st["clean_total"] += total
                left = 0

                for d in dialogs:
                    if st["clean_stop"]: break
                    entity = d.entity
                    skip   = False

                    if isinstance(entity, Channel):
                        if entity.megagroup and not groups:   skip = True
                        if not entity.megagroup and not channels: skip = True
                    elif isinstance(entity, Chat):
                        if not groups: skip = True
                    else:
                        # User/bot
                        if not bots: skip = True

                    if skip:
                        st["clean_progress"] += 1
                        continue

                    try:
                        if isinstance(entity, Channel):
                            await client(LeaveChannelRequest(entity))
                        elif isinstance(entity, Chat):
                            me = await client.get_me()
                            await client(DeleteChatUserRequest(chat_id=entity.id, user_id=me))
                        left += 1
                        acc_mgr_log(st, "clean_log", "success", f"[{phone[-4:]}] Вышел: {d.name or entity.id}")
                        await asyncio.sleep(delay)
                    except Exception as ex:
                        acc_mgr_log(st, "clean_log", "warn", f"[{phone[-4:]}] {d.name}: {str(ex)[:60]}")
                    st["clean_progress"] += 1

                acc_mgr_log(st, "clean_log", "success", f"[{phone[-4:]}] Готово — вышел из {left} чатов")

            except Exception as ex:
                acc_mgr_log(st, "clean_log", "error", f"[{phone[-4:]}] {str(ex)[:80]}")
            finally:
                if client:
                    try: await client.disconnect()
                    except: pass

        st["clean_running"] = False
        st["clean_stop"]    = False
        acc_mgr_log(st, "clean_log", "success", "═══ Очистка завершена ═══")

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(_run())
    loop.close()


@app.route("/api/accounts/clean", methods=["POST"])
@require_license
def api_clean():
    from flask import g
    st = _uam(g.username)

    if st["clean_running"]:
        return jsonify({"ok": False, "error": "Очистка уже запущена"})
    data = (request.get_json(force=True, silent=True) or {})
    st["_username"] = g.username
    st.update({"clean_running": True, "clean_stop": False,
               "clean_log": [], "clean_progress": 0, "clean_total": 0})
    threading.Thread(
        target=clean_task,
        args=(data.get("phones",[]), data.get("groups",True),
              data.get("channels",True), data.get("bots",False),
              int(data.get("delay",3)), st),
        daemon=True
    ).start()
    return jsonify({"ok": True})


@app.route("/api/accounts/clean_stop", methods=["POST"])
@require_license
def api_clean_stop():
    from flask import g
    st = _uam(g.username)

    st["clean_stop"] = True
    return jsonify({"ok": True})


@app.route("/api/accounts/clean_status")
@require_license
def api_clean_status():
    from flask import g
    st = _uam(g.username)

    return jsonify({
        "running":  st["clean_running"],
        "log":      st["clean_log"],
        "progress": st["clean_progress"],
        "total":    st["clean_total"],
    })


# ── Теги ──────────────────────────────────────────────────────────────────

@app.route("/api/accounts/set_tag", methods=["POST"])
@require_license
def api_set_tag():
    from flask import g
    data  = (request.get_json(force=True, silent=True) or {})
    phone = data.get("phone", "").strip()
    tag   = data.get("tag", "").strip()
    if not phone:
        return jsonify({"ok": False, "error": "Нет телефона"})
    cache_f = Path(get_sessions_dir(g.username)) / f"{phone}.cache"
    try:
        cache = {}
        if cache_f.exists():
            cache = json.loads(cache_f.read_text(encoding="utf-8"))
        cache["tag"] = tag
        cache_f.write_text(json.dumps(cache, ensure_ascii=False))
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)[:80]})


# ── Очередь ────────────────────────────────────────────────────────────────

@app.route("/api/queue")
@require_license
def api_queue():
    from flask import g
    st = _ust(g.username)
    return jsonify({"queue": st["queue"], "total": len(st["queue"])})


@app.route("/api/queue/set", methods=["POST"])
@require_license
def api_queue_set():
    text = ((request.get_json(force=True, silent=True) or {})).get("text", "")
    parsed = []
    for line in text.splitlines():
        line = line.strip()
        if not line: continue
        if "|" in line:
            parts = line.split("|", 1)
            title = parts[0].strip()
            username = parts[1].strip().lstrip("@") if len(parts) > 1 else None
        elif ":" in line:
            parts = line.split(":", 1)
            title = parts[0].strip()
            raw = parts[1].strip().lstrip("@") if len(parts) > 1 else None
            username = raw if raw and re.match(r'^[a-zA-Z][a-zA-Z0-9_]{4,31}$', raw or "") else None
        else:
            title, username = line, None
        if title:
            parsed.append({"title": title, "username": username, "status": "pending"})
    from flask import g
    st = _ust(g.username)
    st["queue"] = parsed
    _ulog(st["log"], "info", f"Загружено {len(parsed)} каналов в очередь")
    return jsonify({"ok": True, "count": len(parsed)})


@app.route("/api/queue/clear", methods=["POST"])
@require_license
def api_queue_clear():
    from flask import g
    _ust(g.username)["queue"] = []
    return jsonify({"ok": True})


# ── Запуск ─────────────────────────────────────────────────────────────────

def run_creation_task(st, username="unknown"):
    async def _run():
        try:
            from telethon import TelegramClient, errors
            from telethon.tl.functions.channels import (
                CreateChannelRequest, EditAdminRequest,
                LeaveChannelRequest, UpdateUsernameRequest
            )
            from telethon.tl.functions.account import CheckUsernameRequest
            from telethon.tl.types import ChatAdminRights
        except ImportError:
            _ulog(st["log"], "error", "Telethon не установлен: pip install telethon")
            st["running"] = False
            return

        admin    = st["admin"]
        delay    = st["delay"]
        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)
        accounts = get_accounts(st.get("_username",""))

        if not accounts:
            _ulog(st["log"], "error", "Нет аккаунтов в папке sessions/")
            st["running"] = False
            return

        queue = st["queue"].copy()
        _ulog(st["log"], "success", f"▶ Старт. Аккаунтов: {len(accounts)}, каналов: {len(queue)}")

        # Используем аккаунты по кругу — подключаем каждый только когда нужен
        acc_list = [a["phone"] for a in accounts]
        acc_idx  = 0
        # Пул каналов на аккаунт
        acc_channels = {a["phone"]: a["channels"] for a in accounts}
        # Пропущенные аккаунты (ошибка авт.)
        skipped = set()

        for i, ch in enumerate(queue):
            if st["stop_flag"]:
                _ulog(st["log"], "warn", "⏹ Остановлено")
                break

            st["progress"] = i

            # Ищем следующий рабочий аккаунт
            phone = None
            for _ in range(len(acc_list)):
                cand = acc_list[acc_idx % len(acc_list)]
                acc_idx += 1
                if cand in skipped:
                    continue
                if acc_channels.get(cand, 0) >= 7:
                    continue
                phone = cand
                break

            if not phone:
                _ulog(st["log"], "error", "Все аккаунты исчерпаны или пропущены")
                break

            title    = ch["title"]
            username = ch.get("username")
            sf       = str(Path(get_sessions_dir(st.get("_username",""))) / phone)

            _ulog(st["log"], "info", f"[{phone[-4:]}] Создаю: «{title}»" + (f" (@{username})" if username else ""))

            client = None
            try:
                # Подключаем только этот аккаунт, только сейчас
                client = TelegramClient(sf, api_id, api_hash)
                await client.connect()

                tg = await client(CreateChannelRequest(
                    title=title, about=state.get("description", ""), megagroup=False))
                channel = tg.chats[0]

                assigned_username = None
                if username:
                    try:
                        avail = bool(await client(CheckUsernameRequest(username=username)))
                        if avail:
                            await client(UpdateUsernameRequest(channel, username))
                            assigned_username = username
                        else:
                            _ulog(st["log"], "warn", f"@{username} занят — создаю без username")
                    except Exception as ue:
                        _ulog(st["log"], "warn", f"Username ошибка: {ue}")

                admin_added = creator_left = avatar_set = False
                if admin:
                    try:
                        admin_entity = await client.get_entity(admin.lstrip("@"))
                        rights = ChatAdminRights(
                            change_info=True, post_messages=True, edit_messages=True,
                            delete_messages=True, ban_users=True, invite_users=True,
                            pin_messages=True, add_admins=True)
                        await client(EditAdminRequest(channel, admin_entity, rights, "Администратор"))
                        admin_added = True
                    except Exception as ae:
                        _ulog(st["log"], "warn", f"Админ: {ae}")

                # Ставим рандомную аватарку если есть
                avatar_files = sorted(Path(CHAN_AVATARS_DIR).glob("*"))
                if avatar_files:
                    try:
                        import random
                        from telethon.tl.functions.channels import EditPhotoRequest
                        from telethon.tl.functions.photos import UploadProfilePhotoRequest
                        chosen = str(random.choice(avatar_files))
                        uploaded = await client.upload_file(chosen)
                        await client(EditPhotoRequest(
                            channel=channel,
                            photo=await client.upload_file(chosen)
                        ))
                        avatar_set = True
                        _ulog(st["log"], "success", f"✓ Аватарка установлена: {Path(chosen).name}")
                    except Exception as ae:
                        _ulog(st["log"], "warn", f"Аватарка: {ae}")

                await asyncio.sleep(2)
                if state.get("leave_creator", True):
                    try:
                        await client(LeaveChannelRequest(channel))
                        creator_left = True
                    except Exception as le:
                        _ulog(st["log"], "error", f"LeaveChannel: {le}")

                link = f"t.me/{assigned_username}" if assigned_username else "без username"
                _ulog(st["log"], "success", f"✓ «{title}» — {link} | Админ:{'✓' if admin_added else '✗'} | Аватар:{'✓' if avatar_set else '—'} | Вышел:{'✓' if creator_left else '✗'}")

                ch["status"] = "done"
                acc_channels[phone] = acc_channels.get(phone, 0) + 1
                result = {
                    "id": i+1, "title": title, "username": assigned_username,
                    "phone": phone, "admin": admin,
                    "admin_added": admin_added, "creator_left": creator_left,
                    "success": True, "error": None,
                    "time": datetime.now().strftime("%H:%M:%S")
                }

            except (errors.AuthKeyUnregisteredError, errors.UserDeactivatedError,
                    errors.SessionExpiredError):
                _ulog(st["log"], "error", f"[{phone[-4:]}] Сессия недействительна — пропускаю аккаунт")
                skipped.add(phone)
                ch["status"] = "error"
                result = {
                    "id": i+1, "title": title, "username": username, "phone": phone,
                    "success": False, "error": "Сессия недействительна",
                    "time": datetime.now().strftime("%H:%M:%S")
                }
                # Возвращаем канал в очередь если есть другие аккаунты
                remaining = [p for p in acc_list if p not in skipped]
                if remaining:
                    queue.append(ch)
                    st["total"] += 1
                    continue

            except errors.FloodWaitError as fe:
                _ulog(st["log"], "warn", f"[{phone[-4:]}] FloodWait {fe.seconds}s — жду...")
                ch["status"] = "pending"
                result = {
                    "id": i+1, "title": title, "username": username, "phone": phone,
                    "success": False, "error": f"FloodWait {fe.seconds}s",
                    "time": datetime.now().strftime("%H:%M:%S")
                }
                waited = 0
                while waited < fe.seconds:
                    if st["stop_flag"]: break
                    await asyncio.sleep(min(5, fe.seconds - waited))
                    waited += 5
                # Возвращаем канал в очередь
                if not st["stop_flag"]:
                    queue.append(ch)
                    st["total"] += 1
                    continue

            except Exception as ex:
                _ulog(st["log"], "error", f"✗ «{title}»: {ex}")
                ch["status"] = "error"
                result = {
                    "id": i+1, "title": title, "username": username, "phone": phone,
                    "success": False, "error": str(ex)[:120],
                    "time": datetime.now().strftime("%H:%M:%S")
                }

            finally:
                if client:
                    try: await client.disconnect()
                    except Exception: pass

            st["results"].append(result)
            st["progress"] = i + 1

            if i < len(queue) - 1 and not st["stop_flag"]:
                _ulog(st["log"], "info", f"⏳ Пауза {delay}s...")
                waited = 0
                while waited < delay:
                    if st["stop_flag"]: break
                    await asyncio.sleep(min(5, delay - waited))
                    waited += 5

        ok = sum(1 for r in st["results"] if r["success"])
        _ulog(st["log"], "success", f"✓ Готово. Создано: {ok}/{len(st['queue'])}")

        udir = Path(f"{RESULTS_DIR}/{username}"); udir.mkdir(parents=True, exist_ok=True)
        fname = f"{RESULTS_DIR}/{username}/results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        Path(fname).write_text(json.dumps(st["results"], ensure_ascii=False, indent=2))

        st["running"]   = False
        st["stop_flag"] = False

    async def _safe_run():
        try:
            await _run()
        except Exception as ex:
            _ulog(st["log"], "error", f"Критическая ошибка: {ex}")
        finally:
            st["running"]   = False
            st["stop_flag"] = False

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(_safe_run())
    loop.close()


@app.route("/api/channel_avatars/upload", methods=["POST"])
@require_license
def api_upload_channel_avatars():
    """Загружает несколько аватарок для каналов."""
    files = request.files.getlist("photos")
    if not files:
        return jsonify({"ok": False, "error": "Нет файлов"})
    # Очищаем старые
    for old in Path(CHAN_AVATARS_DIR).glob("*"):
        try: old.unlink()
        except: pass
    saved = []
    for i, f in enumerate(files):
        ext  = Path(f.filename).suffix or ".jpg"
        dest = Path(CHAN_AVATARS_DIR) / f"avatar_{i:03d}{ext}"
        f.save(str(dest))
        saved.append(str(dest))
    return jsonify({"ok": True, "count": len(saved), "files": saved})


@app.route("/api/channel_avatars/list")
@require_license
def api_list_channel_avatars():
    files = sorted(Path(CHAN_AVATARS_DIR).glob("*"))
    return jsonify({"count": len(files), "files": [str(f) for f in files]})


@app.route("/api/channel_avatars/clear", methods=["POST"])
@require_license
def api_clear_channel_avatars():
    for f in Path(CHAN_AVATARS_DIR).glob("*"):
        try: f.unlink()
        except: pass
    return jsonify({"ok": True})


@app.route("/api/run/start", methods=["POST"])
@require_license
def api_run_start():
    from flask import g
    st = _ust(g.username)

    if st["running"]:
        return jsonify({"ok": False, "error": "Уже запущено"}), 400
    if not st["queue"]:
        return jsonify({"ok": False, "error": "Очередь пуста"}), 400
    data = (request.get_json(force=True, silent=True) or {})
    st["admin"]         = data.get("admin", st["admin"]) or ""  # необязательно
    st["delay"]         = int(data.get("delay", st["delay"]))
    st["description"]   = data.get("description", "")
    st["leave_creator"] = bool(data.get("leave_creator", True))
    st["running"]     = True
    st["stop_flag"]   = False
    st["progress"]    = 0
    st["total"]       = len(st["queue"])
    st["results"]     = []
    threading.Thread(target=run_creation_task, args=(st, g.username), daemon=True).start()
    log("info", f"Запущено. Каналов: {st['total']}, задержка: {st['delay']}s")
    return jsonify({"ok": True, "total": st["total"]})


@app.route("/api/run/stop", methods=["POST"])
@require_license
def api_run_stop():
    from flask import g
    st = _ust(g.username)

    st["stop_flag"] = True
    log("warn", "Запрос остановки...")
    return jsonify({"ok": True})


@app.route("/api/run/status")
@require_license
def api_run_status():
    from flask import g
    st = _ust(g.username)

    total = st["total"] or len(st["queue"])
    return jsonify({
        "running": st["running"],
        "progress": st["progress"],
        "total": total,
        "pct": round(st["progress"] / total * 100) if total else 0,
        "success": sum(1 for r in st["results"] if r.get("success")),
        "failed":  sum(1 for r in st["results"] if not r.get("success")),
    })


# ── Лог ───────────────────────────────────────────────────────────────────

@app.route("/api/log")
@require_license
def api_log():
    from flask import g
    st = _ust(g.username)

    since = int(request.args.get("since", 0))
    return jsonify({"log": st["log"][since:], "total": len(st["log"])})


@app.route("/api/log/clear", methods=["POST"])
@require_license
def api_log_clear():
    from flask import g
    st = _ust(g.username)

    st["log"] = []
    log("info", "Лог очищен")
    return jsonify({"ok": True})


# ── Результаты ─────────────────────────────────────────────────────────────

@app.route("/api/results")
@require_license
def api_results():
    from flask import g
    st = _ust(g.username)

    return jsonify({
        "results": st["results"],
        "success": sum(1 for r in st["results"] if r.get("success")),
        "failed":  sum(1 for r in st["results"] if not r.get("success")),
        "total":   len(st["results"]),
    })


@app.route("/api/results/clear", methods=["POST"])
@require_license
def api_results_clear():
    from flask import g
    st = _ust(g.username)

    st["results"] = []
    return jsonify({"ok": True})


@app.route("/api/results/export")
@require_license
def api_results_export():
    from flask import g
    st = _ust(g.username)

    lines = ["#,Название,Username,Аккаунт,Админ,Статус,Время"]
    for r in st["results"]:
        lines.append(f'{r["id"]},"{r["title"]}",{r.get("username") or ""},'
                     f'{r.get("phone","")},{r.get("admin","")},'
                     f'{"OK" if r["success"] else r.get("error","ERR")},'
                     f'{r.get("time","")}')
    return Response("\n".join(lines), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=results_{int(time.time())}.csv"})


# ── Рассылка ───────────────────────────────────────────────────────────────

def bc_log(level, msg):
    entry = {"ts": datetime.now().strftime("%H:%M:%S"), "level": level, "msg": msg}
    st["broadcast_log"].append(entry)
    if len(st["broadcast_log"]) > 500:
        st["broadcast_log"] = st["broadcast_log"][-300:]
    log(level, f"[BC] {msg}")


def broadcast_task(text, recipients, delay, limit, phones_filter, st, buttons=None):
    """
    text       — текст сообщения (HTML)
    recipients — список получателей [@username, +phone, user_id]
    delay      — задержка между отправками (сек)
    limit      — max аккаунтов (0 = все)
    phones_filter — список телефонов отправителей ([] = все)
    buttons    — [[{text, url}, ...], ...] inline keyboard rows
    """
    async def _run():
        try:
            from telethon import TelegramClient, errors
            from telethon.tl.functions.messages import SendMessageRequest
            from telethon.tl.types import (
                ReplyInlineMarkup, KeyboardButtonRow, KeyboardButtonUrl,
                InputPeerUser, InputPeerChannel, InputPeerChat
            )
        except ImportError:
            bc_log("error", "Telethon не установлен — pip install telethon")
            st["broadcast_running"] = False
            return

        # Строим inline keyboard если переданы кнопки
        reply_markup = None
        if buttons:
            try:
                kb_rows = []
                for row in buttons:
                    row_btns = [
                        KeyboardButtonUrl(text=btn["text"], url=btn["url"])
                        for btn in row if btn.get("text") and btn.get("url")
                    ]
                    if row_btns:
                        kb_rows.append(KeyboardButtonRow(buttons=row_btns))
                if kb_rows:
                    reply_markup = ReplyInlineMarkup(rows=kb_rows)
                    bc_log("info", f"Inline keyboard: {sum(len(r.buttons) for r in kb_rows)} кнопок в {len(kb_rows)} строках")
            except Exception as ex:
                bc_log("warn", f"Ошибка создания кнопок: {ex}")

        async def send_msg(client, recipient, msg_text):
            """Отправляет сообщение с кнопками или без."""
            if reply_markup:
                # Парсим HTML вручную — SendMessageRequest не принимает parse_mode
                from telethon.extensions import html as tg_html
                parsed_text, entities = tg_html.parse(msg_text)
                peer = await client.get_input_entity(recipient)
                await client(SendMessageRequest(
                    peer=peer,
                    message=parsed_text,
                    entities=entities or [],
                    reply_markup=reply_markup,
                    no_webpage=True,
                ))
            else:
                await client.send_message(recipient, msg_text, parse_mode='html')

        # Фильтрация аккаунтов-отправителей
        accounts = get_accounts(st.get("_username",""))
        if phones_filter:
            accounts = [a for a in accounts if a["phone"] in phones_filter]
        if limit and limit > 0:
            accounts = accounts[:limit]

        if not accounts:
            bc_log("error", "Нет аккаунтов для рассылки — добавьте сессии")
            st["broadcast_running"] = False
            return

        if not recipients:
            bc_log("error", "Список получателей пуст")
            st["broadcast_running"] = False
            return

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)

        total_ops = len(accounts) * len(recipients)
        st["broadcast_total"] = total_ops
        st["broadcast_sent"]  = 0

        bc_log("info", f"═══ Старт рассылки ═══")
        bc_log("info", f"Аккаунтов: {len(accounts)} | Получателей: {len(recipients)} | Задержка: {delay}s")
        bc_log("info", f"Всего операций: {total_ops}")

        overall_sent = overall_failed = 0

        for acc_idx, acc in enumerate(accounts):
            if st["broadcast_stop"]:
                bc_log("warn", f"⏹ Остановлено. Обработано аккаунтов: {acc_idx}/{len(accounts)}")
                break

            phone = acc["phone"]
            sf = str(Path(get_sessions_dir(st.get("_username",""))) / phone)
            bc_log("info", f"── Аккаунт [{acc_idx+1}/{len(accounts)}]: +{phone} ──")

            client = None
            try:
                client = TelegramClient(sf, api_id, api_hash)
                await client.connect()
                bc_log("info", f"[{phone[-4:]}] Подключён")
            except Exception as conn_ex:
                bc_log("error", f"[{phone[-4:]}] Ошибка подключения: {conn_ex}")
                overall_failed += len(recipients)
                st["broadcast_sent"] += len(recipients)
                continue

            acc_sent = acc_failed = 0

            for rec_idx, recipient in enumerate(recipients):
                if st["broadcast_stop"]:
                    bc_log("warn", f"⏹ Остановлено")
                    break

                # Нормализуем получателя
                r_raw = recipient.strip()
                # username: убираем @ → Telethon принимает без @
                # phone: оставляем +7xxx как есть
                # user_id (число): оставляем как есть
                if r_raw.startswith("@"):
                    clean_recipient = r_raw[1:]  # убираем только @
                else:
                    clean_recipient = r_raw  # phone/id без изменений
                try:
                    await send_msg(client, clean_recipient, text)
                    bc_log("success", f"[{phone[-4:]}] ✓ → {r_raw}")
                    acc_sent += 1
                    overall_sent += 1

                except errors.FloodWaitError as fe:
                    bc_log("warn", f"[{phone[-4:]}] FloodWait {fe.seconds}s на {r_raw} — жду...")
                    waited = 0
                    while waited < fe.seconds:
                        if st["broadcast_stop"]: break
                        await asyncio.sleep(min(5, fe.seconds - waited))
                        waited += 5
                    if not st["broadcast_stop"]:
                        try:
                            await send_msg(client, clean_recipient, text)
                            bc_log("success", f"[{phone[-4:]}] ✓ → {r_raw} (после FloodWait)")
                            acc_sent += 1
                            overall_sent += 1
                        except Exception as retry_ex:
                            bc_log("error", f"[{phone[-4:]}] ✗ → {recipient}: {retry_ex}")
                            acc_failed += 1
                            overall_failed += 1

                except (errors.UsernameInvalidError, errors.UsernameNotOccupiedError):
                    bc_log("warn", f"[{phone[-4:]}] ✗ {r_raw} — пользователь не найден")
                    acc_failed += 1
                    overall_failed += 1

                except errors.PeerFloodError:
                    bc_log("error", f"[{phone[-4:]}] PeerFlood — аккаунт временно ограничен, пропускаю оставшихся")
                    overall_failed += len(recipients) - rec_idx - 1
                    st["broadcast_sent"] += len(recipients) - rec_idx - 1
                    break

                except errors.UserPrivacyRestrictedError:
                    bc_log("warn", f"[{phone[-4:]}] ✗ {r_raw} — приватность запрещает сообщения")
                    acc_failed += 1
                    overall_failed += 1

                except errors.InputUserDeactivatedError:
                    bc_log("warn", f"[{phone[-4:]}] ✗ {r_raw} — аккаунт удалён/деактивирован")
                    acc_failed += 1
                    overall_failed += 1

                except errors.ChatWriteForbiddenError:
                    bc_log("error", f"[{phone[-4:]}] ✗ {r_raw} — нет прав на отправку")
                    acc_failed += 1
                    overall_failed += 1

                except errors.AuthKeyUnregisteredError:
                    bc_log("error", f"[{phone[-4:]}] Сессия истекла/заблокирована — пропуск аккаунта")
                    overall_failed += len(recipients) - rec_idx
                    st["broadcast_sent"] += len(recipients) - rec_idx
                    break

                except errors.SessionExpiredError:
                    bc_log("error", f"[{phone[-4:]}] Сессия устарела — пропуск аккаунта")
                    overall_failed += len(recipients) - rec_idx
                    st["broadcast_sent"] += len(recipients) - rec_idx
                    break

                except Exception as ex:
                    err_msg = str(ex)
                    bc_log("error", f"[{phone[-4:]}] ✗ → {r_raw}: {err_msg[:100]}")
                    acc_failed += 1
                    overall_failed += 1

                st["broadcast_sent"] = overall_sent
                if rec_idx < len(recipients) - 1 and not st["broadcast_stop"]:
                    await asyncio.sleep(delay)

            bc_log("info", f"[{phone[-4:]}] Итог: отправлено {acc_sent}, ошибок {acc_failed}")

            try:
                await client.disconnect()
            except Exception:
                pass

        st["broadcast_runs"]    += 1
        st["broadcast_ok"]      += overall_sent
        st["broadcast_fail"]    += overall_failed
        bc_log("success", f"═══ Рассылка завершена ═══")
        bc_log("success", f"Отправлено: {overall_sent} | Ошибок: {overall_failed} | Всего: {total_ops}")

    async def _safe_run():
        try:
            await _run()
        except Exception as fatal_ex:
            bc_log("error", f"Критическая ошибка: {fatal_ex}")
        finally:
            st["broadcast_running"] = False
            st["broadcast_stop"]    = False

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(_safe_run())
    loop.close()


@app.route("/api/broadcast/send", methods=["POST"])
@require_license
def api_broadcast_send():
    from flask import g
    st = _ust(g.username)

    if st["broadcast_running"]:
        return jsonify({"ok": False, "error": "Рассылка уже запущена"}), 400

    data       = (request.get_json(force=True, silent=True) or {})
    text       = data.get("text", "").strip()
    recipients = data.get("recipients", [])
    delay      = max(1, int(data.get("delay", 4)))
    limit      = int(data.get("limit", 0))
    phones     = data.get("phones", [])
    buttons    = data.get("buttons", [])   # [[{text,url},...],...]

    if not text:
        return jsonify({"ok": False, "error": "Текст сообщения пуст"}), 400
    if not recipients:
        return jsonify({"ok": False, "error": "Список получателей пуст"}), 400

    all_accs = get_accounts(g.username)
    filtered = [a for a in all_accs if not phones or a["phone"] in phones]
    if limit and limit > 0:
        filtered = filtered[:limit]
    if not filtered:
        return jsonify({"ok": False, "error": "Нет доступных аккаунтов (добавьте сессии)"}), 400

    total_ops = len(filtered) * len(recipients)
    st["_username"] = g.username
    st["broadcast_running"] = True
    st["broadcast_stop"]    = False
    st["broadcast_sent"]    = 0
    st["broadcast_total"]   = total_ops

    threading.Thread(
        target=broadcast_task,
        args=(text, recipients, delay, limit, phones, st),
        kwargs={"buttons": buttons},
        daemon=True
    ).start()

    return jsonify({
        "ok": True,
        "accounts": len(filtered),
        "recipients": len(recipients),
        "total": total_ops,
    })


@app.route("/api/broadcast/stop", methods=["POST"])
@require_license
def api_broadcast_stop():
    from flask import g
    st = _ust(g.username)

    st["broadcast_stop"] = True
    bc_log("warn", "Запрос остановки рассылки...")
    return jsonify({"ok": True})


@app.route("/api/broadcast/log")
@require_license
def api_broadcast_log():
    from flask import g
    st = _ust(g.username)

    since = int(request.args.get("since", 0))
    return jsonify({"log": st["broadcast_log"][since:], "total": len(st["broadcast_log"])})


@app.route("/api/broadcast/stats")
@require_license
def api_broadcast_stats():
    from flask import g
    st = _ust(g.username)

    return jsonify({
        "runs":    st["broadcast_runs"],
        "ok":      st["broadcast_ok"],
        "fail":    st["broadcast_fail"],
        "running": st["broadcast_running"],
        "sent":    st["broadcast_sent"],
        "total":   st["broadcast_total"],
    })


# ── Парсер аудитории ──────────────────────────────────────────────────────

parser_state = {
    "running": False, "stop_flag": False,
    "progress": 0, "total": 0,
    "found": 0, "filtered": 0,
    "log": [], "results": [],
}

PARSER_RESULTS_DIR = "parser_results"
BASES_DIR          = "parser_bases"
Path(BASES_DIR).mkdir(exist_ok=True)
Path(PARSER_RESULTS_DIR).mkdir(exist_ok=True)


def parser_log(st, level, msg):
    _ulog(st["log"], level, msg, "PARSER")


def _gender_guess(first_name: str) -> str:
    """Простая эвристика по имени для определения пола."""
    if not first_name:
        return ""
    name = first_name.lower().strip()
    female_endings = ("а", "я", "ия", "ья", "ина", "на")
    male_endings   = ("й", "ий", "ей", "ев", "ов", "ин", "ен", "он", "ан")
    for e in female_endings:
        if name.endswith(e):
            return "female"
    for e in male_endings:
        if name.endswith(e):
            return "male"
    return ""


def _last_online_str(status) -> str:
    try:
        from telethon.tl.types import (
            UserStatusOnline, UserStatusOffline, UserStatusRecently,
            UserStatusLastWeek, UserStatusLastMonth,
        )
        if isinstance(status, UserStatusOnline):
            return "онлайн"
        if isinstance(status, UserStatusRecently):
            return "недавно"
        if isinstance(status, UserStatusLastWeek):
            return "неделю назад"
        if isinstance(status, UserStatusLastMonth):
            return "месяц назад"
        if isinstance(status, UserStatusOffline) and status.was_online:
            return status.was_online.strftime("%d.%m.%Y %H:%M")
    except Exception:
        pass
    return "—"


def _passes_filters(user_data: dict, filters: dict) -> bool:
    """Проверяет соответствие пользователя фильтрам."""
    from datetime import datetime, timezone, timedelta

    # Фильтр пола
    if filters.get("gender"):
        if user_data.get("gender") != filters["gender"]:
            return False

    # Фильтр аватарки
    if filters.get("avatar") == "yes" and not user_data.get("has_avatar"):
        return False
    if filters.get("avatar") == "no" and user_data.get("has_avatar"):
        return False

    # Фильтр Premium
    if filters.get("premium") == "yes" and not user_data.get("is_premium"):
        return False
    if filters.get("premium") == "no" and user_data.get("is_premium"):
        return False

    # Фильтр юзернейма
    if filters.get("username") == "yes" and not user_data.get("username"):
        return False
    if filters.get("username") == "no" and user_data.get("username"):
        return False

    # Фильтр онлайна
    online_filter = filters.get("online", "")
    if online_filter:
        raw_online = user_data.get("_last_online_dt")
        if raw_online is None:
            return False
        now = datetime.now(timezone.utc)
        deltas = {"1d": 1, "3d": 3, "7d": 7, "30d": 30, "month": 30}
        days = deltas.get(online_filter, 0)
        if days:
            if (now - raw_online).days > days:
                return False

    return True


def _extract_user(user, source: str) -> dict:
    """Извлекает данные пользователя в dict."""
    from telethon.tl.types import (
        UserStatusOnline, UserStatusOffline, UserStatusRecently,
        UserStatusLastWeek, UserStatusLastMonth,
    )
    from datetime import datetime, timezone, timedelta

    gender     = _gender_guess(user.first_name or "")
    has_avatar = bool(user.photo)
    # Правильный атрибут premium в Telethon
    is_premium = bool(getattr(user, "premium", False))

    last_online_dt = None
    try:
        if isinstance(user.status, UserStatusOnline):
            last_online_dt = datetime.now(timezone.utc)
        elif isinstance(user.status, UserStatusOffline) and user.status.was_online:
            wo = user.status.was_online
            last_online_dt = wo if wo.tzinfo else wo.replace(tzinfo=timezone.utc)
        elif isinstance(user.status, UserStatusRecently):
            last_online_dt = datetime.now(timezone.utc) - timedelta(hours=12)
        elif isinstance(user.status, UserStatusLastWeek):
            last_online_dt = datetime.now(timezone.utc) - timedelta(days=5)
        elif isinstance(user.status, UserStatusLastMonth):
            last_online_dt = datetime.now(timezone.utc) - timedelta(days=20)
    except Exception:
        pass

    return {
        "id":              user.id,
        "first_name":      user.first_name or "",
        "last_name":       user.last_name or "",
        "username":        user.username or "",
        "phone":           getattr(user, "phone", "") or "",
        "gender":          gender,
        "has_avatar":      has_avatar,
        "is_premium":      is_premium,
        "last_online":     _last_online_str(user.status),
        "_last_online_dt": last_online_dt,
        "source":          source,
    }


def parser_task(sources, mode, filters, limit, phones, st):
    async def _run():
        try:
            from telethon import TelegramClient, errors
            from telethon.tl.functions.messages import GetHistoryRequest
        except ImportError:
            parser_log(st, "error", "Telethon не установлен: pip install telethon")
            st["running"] = False
            return

        accounts = get_accounts(st.get("_username",""))
        if phones:
            accounts = [a for a in accounts if a["phone"] in phones]
        if not accounts:
            parser_log(st, "error", "Нет аккаунтов")
            st["running"] = False
            return

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)

        all_users: dict[int, dict] = {}
        st["total"]    = len(sources)
        st["progress"] = 0
        st["found"]    = 0
        st["filtered"] = 0
        st["results"]  = []

        # Распределяем аккаунты по источникам
        acc_idx = 0

        for src_idx, source in enumerate(sources):
            if st["stop_flag"]:
                parser_log(st, "warn", "⏹ Остановлено")
                break

            parser_log(st, "info", f"[{src_idx+1}/{len(sources)}] Парсинг: {source} ({mode})")
            phone = accounts[acc_idx % len(accounts)]["phone"]
            acc_idx += 1
            session_file = str(Path(get_sessions_dir(st.get("_username",""))) / phone)

            client = None
            try:
                client = TelegramClient(session_file, api_id, api_hash)
                await client.connect()
                if not await client.is_user_authorized():
                    parser_log(st, "warn", f"[{phone[-4:]}] Не авторизован, пропуск")
                    continue

                entity = await client.get_entity(source.lstrip("@").replace("https://t.me/",""))
                src_users_raw = []

                def is_real_user(obj):
                    """Возвращает True только если obj — реальный пользователь (не бот, не канал)."""
                    from telethon.tl.types import User
                    return obj is not None and isinstance(obj, User) and not obj.bot

                if mode == "members":
                    async for u in client.iter_participants(entity):
                        if st["stop_flag"]: break
                        if is_real_user(u):
                            src_users_raw.append(u)

                elif mode == "commenters":
                    seen_comm = set()
                    async for msg in client.iter_messages(entity, limit=200):
                        if st["stop_flag"]: break
                        if msg.replies and msg.replies.replies > 0:
                            try:
                                async for reply in client.iter_messages(entity, reply_to=msg.id, limit=100):
                                    if is_real_user(reply.sender) and reply.sender.id not in seen_comm:
                                        seen_comm.add(reply.sender.id)
                                        src_users_raw.append(reply.sender)
                            except Exception:
                                pass

                elif mode == "messages":
                    seen = set()
                    async for msg in client.iter_messages(entity, limit=3000):
                        if st["stop_flag"]: break
                        if is_real_user(msg.sender) and msg.sender.id not in seen:
                            seen.add(msg.sender.id)
                            src_users_raw.append(msg.sender)

                elif mode == "reactions":
                    seen_react = set()
                    async for msg in client.iter_messages(entity, limit=100):
                        if st["stop_flag"]: break
                        if not msg.reactions:
                            continue
                        try:
                            from telethon.tl.functions.messages import GetMessageReactionsListRequest
                            offset = ""
                            while True:
                                kwargs = {"peer": entity, "id": msg.id, "limit": 100}
                                if offset:
                                    kwargs["offset"] = offset
                                try:
                                    res = await client(GetMessageReactionsListRequest(**kwargs))
                                except Exception:
                                    break
                                for reaction in res.reactions:
                                    try:
                                        u = await client.get_entity(reaction.peer_id)
                                        if is_real_user(u) and u.id not in seen_react:
                                            seen_react.add(u.id)
                                            src_users_raw.append(u)
                                    except Exception:
                                        pass
                                if not res.next_offset:
                                    break
                                offset = res.next_offset
                        except Exception:
                            pass

                elif mode == "polls":
                    async for msg in client.iter_messages(entity, limit=200):
                        if st["stop_flag"]: break
                        if msg.poll:
                            try:
                                from telethon.tl.functions.messages import GetPollVotersRequest
                                for opt in msg.poll.poll.answers:
                                    voters = await client(GetPollVotersRequest(peer=entity, id=msg.id, option=opt.option, limit=200))
                                    for v in voters.users:
                                        if is_real_user(v):
                                            src_users_raw.append(v)
                            except Exception:
                                pass

                elif mode == "join":
                    seen_join = set()
                    async for msg in client.iter_messages(entity, limit=5000):
                        if st["stop_flag"]: break
                        from telethon.tl.types import MessageActionChatAddUser, MessageActionChatJoinedByLink
                        if msg.action and isinstance(msg.action, (MessageActionChatAddUser, MessageActionChatJoinedByLink)):
                            if is_real_user(msg.sender) and msg.sender.id not in seen_join:
                                seen_join.add(msg.sender.id)
                                src_users_raw.append(msg.sender)

                # Конвертируем и фильтруем
                new_count = 0
                for u in src_users_raw:
                    if u.id in all_users:
                        continue
                    ud = _extract_user(u, source)
                    all_users[u.id] = ud
                    st["found"] += 1
                    new_count += 1

                parser_log(st, "success", f"✓ {source} — новых пользователей: {new_count}")

            except errors.FloodWaitError as fe:
                parser_log(st, "warn", f"FloodWait {fe.seconds}s...")
                await asyncio.sleep(min(fe.seconds, 60))
            except Exception as ex:
                parser_log(st, "error", f"✗ {source}: {str(ex)[:120]}")
            finally:
                if client:
                    try: await client.disconnect()
                    except Exception: pass

            st["progress"] = src_idx + 1
            await asyncio.sleep(2)

        # Применяем фильтры
        parser_log(st, "info", f"Применяю фильтры к {len(all_users)} пользователям...")
        filtered = [u for u in all_users.values() if _passes_filters(u, filters)]
        if limit and limit > 0:
            filtered = filtered[:limit]

        # Убираем служебные поля
        for u in filtered:
            u.pop("_last_online_dt", None)

        st["filtered"] = len(filtered)
        st["results"]  = filtered
        st["running"]  = False
        st["stop_flag"] = False
        parser_log(st, "success", f"═══ Готово. Всего: {st['found']}, после фильтра: {len(filtered)} ═══")

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(_run())
    loop.close()


@app.route("/api/parser/start", methods=["POST"])
@require_license
def api_parser_start():
    from flask import g
    st = _uprs(g.username)

    if st["running"]:
        return jsonify({"ok": False, "error": "Парсер уже запущен"}), 400
    data    = (request.get_json(force=True, silent=True) or {})
    sources = data.get("sources", [])
    mode    = data.get("mode", "members")
    filters = data.get("filters", {})
    limit   = int(data.get("limit", 0))
    phones  = data.get("phones", [])
    if not sources:
        return jsonify({"ok": False, "error": "Нет источников"}), 400

    st["running"]   = True
    st["stop_flag"] = False
    st["log"]       = []
    st["results"]   = []
    st["found"]     = 0
    st["filtered"]  = 0
    st["progress"]  = 0
    st["total"]     = len(sources)

    st["_username"] = g.username
    threading.Thread(
        target=parser_task,
        args=(sources, mode, filters, limit, phones, st),
        daemon=True
    ).start()
    return jsonify({"ok": True})


@app.route("/api/parser/stop", methods=["POST"])
@require_license
def api_parser_stop():
    from flask import g
    st = _uprs(g.username)

    st["stop_flag"] = True
    parser_log(st, "warn", "Запрос остановки...")
    return jsonify({"ok": True})


@app.route("/api/parser/status")
@require_license
def api_parser_status():
    from flask import g
    st = _uprs(g.username)

    return jsonify({
        "running":  st["running"],
        "progress": st["progress"],
        "total":    st["total"],
        "found":    st["found"],
        "filtered": st["filtered"],
        "log":      st["log"],
        "results":  st["results"],
    })


@app.route("/api/parser/export", methods=["POST"])
@require_license
def api_parser_export():
    from flask import g
    st = _uprs(g.username)

    data    = (request.get_json(force=True, silent=True) or {})
    results = data.get("results", st["results"])
    if not results:
        return jsonify({"ok": False, "error": "Нет данных"}), 400
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"ok": False, "error": "openpyxl не установлен: pip install openpyxl"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Пользователи"

    headers = ["#","Имя","Фамилия","Юзернейм","ID","Телефон","Пол","Последний онлайн","Аватарка","Premium","Источник"]
    header_fill = PatternFill("solid", fgColor="1a1e27")
    header_font = Font(bold=True, color="4ade80")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, u in enumerate(results, 1):
        ws.append([
            i,
            u.get("first_name",""),
            u.get("last_name",""),
            "@"+u["username"] if u.get("username") else "",
            u.get("id",""),
            u.get("phone",""),
            "М" if u.get("gender")=="male" else ("Ж" if u.get("gender")=="female" else ""),
            u.get("last_online",""),
            "Да" if u.get("has_avatar") else "Нет",
            "Да" if u.get("is_premium") else "Нет",
            u.get("source",""),
        ])

    # Авторазмер столбцов
    col_widths = [4,18,18,20,14,16,6,20,10,10,26]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"parsed_{ts_str}.xlsx"
    filepath = Path(PARSER_RESULTS_DIR) / filename
    wb.save(str(filepath))
    parser_log(st, "success", f"Excel сохранён: {filename} ({len(results)} строк)")

    return jsonify({"ok": True, "url": f"/parser_results/{filename}", "filename": filename})


@app.route("/parser_results/<path:filename>")
@require_license
def serve_parser_result(filename):
    return send_from_directory(PARSER_RESULTS_DIR, filename, as_attachment=True)


# ── Инвайтер ──────────────────────────────────────────────────────────────

inviter_state = {
    "running": False, "stop_flag": False,
    "progress": 0, "total": 0,
    "ok": 0, "fail": 0, "swaps": 0,
    "log": [], "chat_stats": {},
    "cur_acc": "", "cur_chat": "",
    # Фаза подготовки
    "prepare_running": False, "prepare_done": False,
    "prepare_phase": "",        # "join" | "grant"
    "prepare_progress": 0, "prepare_total": 0,
    "prepare_detail": "",
}


def inv_log(st, level, msg):
    _ulog(st["log"], level, msg, "INV")


def inviter_task(chats, users, mode, per_acc, swap_after, delay, flood_delay,
                 distribute, autoswap, phones,
                 prepare=False, master_acc="", join_delay=8, grant_delay=5, st=None):
    async def _run():
        try:
            from telethon import TelegramClient, errors
            from telethon.tl.functions.channels import (
                InviteToChannelRequest, EditAdminRequest,
                JoinChannelRequest, EditBannedRequest
            )
            from telethon.tl.functions.messages import AddChatUserRequest
            from telethon.tl.types import (
                Channel, Chat, ChatAdminRights, ChatBannedRights
            )
        except ImportError:
            inv_log(st, "error", "Telethon не установлен: pip install telethon")
            st["running"] = False
            return

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)

        accounts = get_accounts(st.get("_username","") if st else "")
        if phones:
            accounts = [a for a in accounts if a["phone"] in phones]
        if not accounts:
            inv_log(st, "error", "Нет аккаунтов для работы")
            st["running"] = False
            return

        inv_log(st, "info", f"Аккаунтов: {len(accounts)}, чатов: {len(chats)}, польз.: {len(users)}")
        for c in chats:
            st["chat_stats"][c] = {"ok": 0, "fail": 0}

        # ─── ФАЗА 1: Вступление в чаты ───────────────────────────────────────
        if prepare and master_acc:
            worker_phones = [a["phone"] for a in accounts if a["phone"] != master_acc]
            if not worker_phones:
                inv_log(st, "warn", "Нет рабочих аккаунтов кроме главного — пропуск подготовки")
            else:
                total_ops = len(worker_phones) * len(chats)
                if st: st.update({
                    "prepare_running": True, "prepare_done": False,
                    "prepare_phase": "join", "prepare_total": total_ops,
                    "prepare_progress": 0, "prepare_detail": ""
                })
                inv_log(st, "info", f"=== ФАЗА 1: Вступление ({len(worker_phones)} акк x {len(chats)} чатов) ===")

                op = 0
                for phone in worker_phones:
                    if st["stop_flag"]: break
                    client = None
                    try:
                        client = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / phone), api_id, api_hash)
                        await client.connect()
                        if not await client.is_user_authorized():
                            inv_log(st, "warn", f"[{phone[-4:]}] Не авторизован, пропуск")
                            op += len(chats)
                            st["prepare_progress"] = op
                            continue
                        for chat_ref in chats:
                            if st["stop_flag"]: break
                            st["prepare_detail"] = f"[{phone[-4:]}] -> {chat_ref}"
                            try:
                                chat_slug = chat_ref.lstrip("@").replace("https://t.me/", "")
                                entity = await client.get_entity(chat_slug)
                                already_in = False
                                try:
                                    await client.get_permissions(entity, await client.get_me())
                                    already_in = True
                                except Exception:
                                    pass
                                if already_in:
                                    inv_log(st, "info", f"[{phone[-4:]}] Уже в {chat_ref}")
                                else:
                                    await client(JoinChannelRequest(entity))
                                    inv_log(st, "success", f"[{phone[-4:]}] Вступил в {chat_ref}")
                                    await asyncio.sleep(join_delay)
                            except errors.FloodWaitError as fe:
                                inv_log(st, "warn", f"[{phone[-4:]}] FloodWait {fe.seconds}s")
                                await asyncio.sleep(min(fe.seconds, 60))
                            except Exception as ex:
                                inv_log(st, "warn", f"[{phone[-4:]}] Вступление в {chat_ref}: {str(ex)[:80]}")
                            op += 1
                            st["prepare_progress"] = op
                    except Exception as ex:
                        inv_log(st, "error", f"[{phone[-4:]}] Ошибка: {str(ex)[:80]}")
                        op += len(chats)
                        st["prepare_progress"] = op
                    finally:
                        if client:
                            try: await client.disconnect()
                            except: pass

                # ─── ФАЗА 2: Выдача прав ─────────────────────────────────────
                inv_log(st, "info", f"=== ФАЗА 2: Выдача прав (главный: {master_acc[-4:]}) ===")
                inviter_state.update({
                    "prepare_phase": "grant",
                    "prepare_total": len(worker_phones) * len(chats),
                    "prepare_progress": 0
                })
                rights = ChatAdminRights(
                    invite_users=True, add_admins=True, change_info=True,
                    post_messages=True, edit_messages=True, delete_messages=True,
                    ban_users=True, pin_messages=True,
                )
                master_client = None
                try:
                    master_client = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / master_acc), api_id, api_hash)
                    await master_client.connect()
                    if not await master_client.is_user_authorized():
                        inv_log(st, "error", f"Главный [{master_acc[-4:]}] не авторизован!")
                    else:
                        worker_id_map = {}
                        for phone in worker_phones:
                            try:
                                tmp = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / phone), api_id, api_hash)
                                await tmp.connect()
                                me = await tmp.get_me()
                                await tmp.disconnect()
                                if me:
                                    worker_id_map[me.id] = phone
                            except Exception:
                                pass

                        op = 0
                        for chat_ref in chats:
                            if st["stop_flag"]: break
                            try:
                                chat_slug = chat_ref.lstrip("@").replace("https://t.me/", "")
                                chat_entity = await master_client.get_entity(chat_slug)
                                members_map = {}
                                async for member in master_client.iter_participants(chat_entity, limit=5000):
                                    if member.id in worker_id_map:
                                        members_map[member.id] = member
                                    if len(members_map) >= len(worker_id_map):
                                        break
                                for uid, phone in worker_id_map.items():
                                    if st["stop_flag"]: break
                                    st["prepare_detail"] = f"[{phone[-4:]}] -> {chat_ref}"
                                    member = members_map.get(uid)
                                    if not member:
                                        inv_log(st, "warn", f"[{phone[-4:]}] Не найден в {chat_ref}")
                                        op += 1
                                        st["prepare_progress"] = op
                                        continue
                                    try:
                                        await master_client(EditAdminRequest(
                                            channel=chat_entity, user_id=member,
                                            admin_rights=rights, rank="Инвайтер"
                                        ))
                                        inv_log(st, "success", f"[{phone[-4:]}] Права выданы в {chat_ref}")
                                        await asyncio.sleep(grant_delay)
                                    except errors.FloodWaitError as fe:
                                        inv_log(st, "warn", f"FloodWait {fe.seconds}s при выдаче прав")
                                        await asyncio.sleep(min(fe.seconds, 30))
                                    except Exception as ex:
                                        inv_log(st, "warn", f"[{phone[-4:]}] Права в {chat_ref}: {str(ex)[:80]}")
                                    op += 1
                                    st["prepare_progress"] = op
                            except Exception as ex:
                                inv_log(st, "error", f"Выдача прав в {chat_ref}: {str(ex)[:80]}")
                except Exception as ex:
                    inv_log(st, "error", f"Ошибка главного клиента: {str(ex)[:100]}")
                finally:
                    if master_client:
                        try: await master_client.disconnect()
                        except: pass

                st["prepare_running"] = False
                st["prepare_done"]    = True
                inv_log(st, "success", "=== Подготовка завершена. Запускаю инвайтинг... ===")
                await asyncio.sleep(2)

        # ─── ОСНОВНОЙ ИНВАЙТИНГ ──────────────────────────────────────────────
        if distribute:
            assignments = [(users[i], chats[i % len(chats)]) for i in range(len(users))]
        else:
            assignments = [(u, c) for c in chats for u in users]

        inviter_state.update({
            "total": len(assignments), "progress": 0,
            "ok": 0, "fail": 0, "swaps": 0,
        })
        inv_log(st, "success", f"Инвайтинг: {len(assignments)} операций, {len(accounts)} аккаунтов")

        acc_idx         = 0
        acc_inv_count   = 0
        acc_flood_count = 0

        def next_account():
            nonlocal acc_idx, acc_inv_count, acc_flood_count
            acc_idx = (acc_idx + 1) % len(accounts)
            acc_inv_count = 0
            acc_flood_count = 0
            st["swaps"] += 1
            inv_log(st, "warn", f"Смена аккаунта -> [{accounts[acc_idx]['phone'][-4:]}]")

        for idx, (user_ref, chat_ref) in enumerate(assignments):
            if st["stop_flag"]:
                inv_log(st, "warn", "Остановлено")
                break

            if acc_inv_count >= per_acc and autoswap:
                next_account()

            phone = accounts[acc_idx % len(accounts)]["phone"]
            st["cur_acc"]  = phone[-4:]
            st["cur_chat"] = chat_ref

            client = None
            try:
                client = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / phone), api_id, api_hash)
                await client.connect()

                if not await client.is_user_authorized():
                    inv_log(st, "warn", f"[{phone[-4:]}] Не авторизован")
                    if autoswap: next_account()
                    st["progress"] = idx + 1
                    continue

                chat_entity = await client.get_entity(
                    chat_ref.lstrip("@").replace("https://t.me/", "")
                )

                user_ref_clean = user_ref.strip()
                try:
                    if user_ref_clean.lstrip("-").isdigit():
                        user_entity = await client.get_entity(int(user_ref_clean))
                    else:
                        user_entity = await client.get_entity(user_ref_clean.lstrip("@"))
                except Exception as ue:
                    inv_log(st, "warn", f"Не найден {user_ref}: {str(ue)[:60]}")
                    st["fail"] += 1
                    st["chat_stats"].setdefault(chat_ref, {"ok":0,"fail":0})["fail"] += 1
                    st["progress"] = idx + 1
                    continue

                if isinstance(chat_entity, Channel):
                    await client(InviteToChannelRequest(channel=chat_entity, users=[user_entity]))
                else:
                    await client(AddChatUserRequest(chat_id=chat_entity.id, user_id=user_entity, fwd_limit=10))

                inv_log(st, "success", f"+ {user_ref} -> {chat_ref}")
                st["ok"] += 1
                st["chat_stats"].setdefault(chat_ref, {"ok":0,"fail":0})["ok"] += 1
                acc_inv_count += 1
                await asyncio.sleep(delay)

            except errors.FloodWaitError as fe:
                acc_flood_count += 1
                inv_log(st, "warn", f"[{phone[-4:]}] FloodWait {fe.seconds}s (#{acc_flood_count})")
                st["fail"] += 1
                st["chat_stats"].setdefault(chat_ref, {"ok":0,"fail":0})["fail"] += 1
                if autoswap and acc_flood_count >= swap_after:
                    inv_log(st, "warn", f"{swap_after} списаний — смена")
                    next_account()
                else:
                    await asyncio.sleep(min(fe.seconds, flood_delay))

            except errors.UserAlreadyParticipantError:
                inv_log(st, "info", f"· {user_ref} уже в {chat_ref}")

            except errors.UserPrivacyRestrictedError:
                inv_log(st, "warn", f"x {user_ref} — приватность")
                st["fail"] += 1
                st["chat_stats"].setdefault(chat_ref, {"ok":0,"fail":0})["fail"] += 1

            except errors.PeerFloodError:
                inv_log(st, "error", f"[{phone[-4:]}] PeerFlood — смена")
                st["fail"] += 1
                st["chat_stats"].setdefault(chat_ref, {"ok":0,"fail":0})["fail"] += 1
                if autoswap: next_account()
                await asyncio.sleep(flood_delay)

            except (errors.AuthKeyUnregisteredError, errors.SessionExpiredError,
                    errors.UserDeactivatedError):
                inv_log(st, "error", f"[{phone[-4:]}] Сессия мертва — смена")
                if autoswap: next_account()

            except Exception as ex:
                inv_log(st, "error", f"x {user_ref} -> {chat_ref}: {str(ex)[:100]}")
                st["fail"] += 1
                st["chat_stats"].setdefault(chat_ref, {"ok":0,"fail":0})["fail"] += 1

            finally:
                if client:
                    try: await client.disconnect()
                    except: pass

            st["progress"] = idx + 1

        inv_log(st, "success", f"=== Инвайтинг завершён. +{st['ok']} x{st['fail']} ===")

        # ─── ФАЗА 3: Удаление аккаунтов из чатов ────────────────────────────
        if prepare and master_acc:
            inv_log(st, "info", "=== ФАЗА 3: Удаление аккаунтов из чатов ===")
            try:
                worker_id_map2 = {}
                for phone in [a["phone"] for a in accounts if a["phone"] != master_acc]:
                    try:
                        tmp = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / phone), api_id, api_hash)
                        await tmp.connect()
                        me = await tmp.get_me()
                        await tmp.disconnect()
                        if me:
                            worker_id_map2[me.id] = phone
                    except Exception:
                        pass

                mc = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / master_acc), api_id, api_hash)
                await mc.connect()
                try:
                    for chat_ref in chats:
                        try:
                            chat_slug = chat_ref.lstrip("@").replace("https://t.me/", "")
                            chat_entity = await mc.get_entity(chat_slug)
                            members_map2 = {}
                            async for member in mc.iter_participants(chat_entity, limit=5000):
                                if member.id in worker_id_map2:
                                    members_map2[member.id] = member
                                if len(members_map2) >= len(worker_id_map2):
                                    break
                            for uid, phone in worker_id_map2.items():
                                member = members_map2.get(uid)
                                if not member:
                                    continue
                                try:
                                    await mc(EditBannedRequest(
                                        channel=chat_entity, participant=member,
                                        banned_rights=ChatBannedRights(until_date=None, view_messages=True)
                                    ))
                                    await asyncio.sleep(1)
                                    await mc(EditBannedRequest(
                                        channel=chat_entity, participant=member,
                                        banned_rights=ChatBannedRights(until_date=None)
                                    ))
                                    inv_log(st, "success", f"[{phone[-4:]}] Удалён из {chat_ref}")
                                    await asyncio.sleep(2)
                                except Exception as ex:
                                    inv_log(st, "warn", f"[{phone[-4:]}] Удаление из {chat_ref}: {str(ex)[:80]}")
                        except Exception as ex:
                            inv_log(st, "error", f"Очистка {chat_ref}: {str(ex)[:80]}")
                finally:
                    try: await mc.disconnect()
                    except: pass
                inv_log(st, "success", "=== Аккаунты удалены из всех чатов ===")
            except Exception as ex:
                inv_log(st, "error", f"Ошибка очистки: {str(ex)[:100]}")

        st["running"]   = False
        st["stop_flag"] = False

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(_run())
    loop.close()


@app.route("/api/inviter/start", methods=["POST"])
@require_license
def api_inviter_start():
    from flask import g
    st = _uinv(g.username)

    if st["running"]:
        return jsonify({"ok": False, "error": "Инвайтер уже запущен"}), 400
    data = (request.get_json(force=True, silent=True) or {})
    chats  = data.get("chats", [])
    users  = data.get("users", [])
    if not chats: return jsonify({"ok": False, "error": "Нет чатов"}), 400
    if not users: return jsonify({"ok": False, "error": "Нет пользователей"}), 400
    if len(chats) > 100: return jsonify({"ok": False, "error": "Максимум 100 чатов"}), 400

    prepare    = bool(data.get("prepare", False))
    master_acc = data.get("master_acc", "").strip()
    if prepare and not master_acc:
        return jsonify({"ok": False, "error": "Укажите главного администратора"}), 400

    st["_username"] = g.username
    st.update({
        "running": True, "stop_flag": False,
        "progress": 0, "total": 0,
        "ok": 0, "fail": 0, "swaps": 0,
        "log": [], "chat_stats": {},
        "cur_acc": "", "cur_chat": "",
        "prepare_running": False, "prepare_done": False,
        "prepare_phase": "", "prepare_progress": 0,
        "prepare_total": 0, "prepare_detail": "",
    })

    threading.Thread(
        target=inviter_task,
        kwargs=dict(
            st=st,
            chats=chats, users=users,
            mode=data.get("mode", "add"),
            per_acc=int(data.get("per_acc", 30)),
            swap_after=int(data.get("swap_after", 5)),
            delay=int(data.get("delay", 15)),
            flood_delay=int(data.get("flood_delay", 60)),
            distribute=bool(data.get("distribute", True)),
            autoswap=bool(data.get("autoswap", True)),
            phones=data.get("phones", []),
            prepare=prepare,
            master_acc=master_acc,
            join_delay=int(data.get("join_delay", 8)),
            grant_delay=int(data.get("grant_delay", 5)),
        ),
        daemon=True
    ).start()
    return jsonify({"ok": True})


@app.route("/api/inviter/stop", methods=["POST"])
@require_license
def api_inviter_stop():
    from flask import g
    st = _uinv(g.username)

    st["stop_flag"] = True
    inv_log(st, "warn", "Запрос остановки инвайтера...")
    return jsonify({"ok": True})


@app.route("/api/inviter/status")
@require_license
def api_inviter_status():
    from flask import g
    st = _uinv(g.username)

    return jsonify({
        "running":          st["running"],
        "progress":         st["progress"],
        "total":            st["total"],
        "ok":               st["ok"],
        "fail":             st["fail"],
        "swaps":            st["swaps"],
        "cur_acc":          st["cur_acc"],
        "cur_chat":         st["cur_chat"],
        "chat_stats":       st["chat_stats"],
        "log":              st["log"],
        "prepare_running":  st["prepare_running"],
        "prepare_done":     st["prepare_done"],
        "prepare_phase":    st["prepare_phase"],
        "prepare_progress": st["prepare_progress"],
        "prepare_total":    st["prepare_total"],
        "prepare_detail":   st["prepare_detail"],
    })


# ── Чекер номеров ─────────────────────────────────────────────────────────

pc_state = {
    "running": False, "stop_flag": False,
    "progress": 0, "total": 0,
    "found": 0, "notfound": 0, "errors": 0,
    "log": [], "results": [], "cur": "",
}


def pc_log(st, level, msg):
    _ulog(st["log"], level, msg, "PC")


def phonechecker_task(phones, acc_phones, delay, st=None):
    async def _run():
        try:
            from telethon import TelegramClient, errors
            from telethon.tl.functions.contacts import ImportContactsRequest, DeleteContactsRequest
            from telethon.tl.types import InputPhoneContact
        except ImportError:
            pc_log(st, "error", "Telethon не установлен")
            st["running"] = False
            return

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)
        accounts = get_accounts(st.get("_username",""))
        if acc_phones:
            accounts = [a for a in accounts if a["phone"] in acc_phones]
        if not accounts:
            pc_log(st, "error", "Нет аккаунтов для проверки")
            st["running"] = False
            return

        # Используем первый доступный аккаунт (или чередуем)
        phone     = accounts[0]["phone"]
        sf        = str(Path(get_sessions_dir(st.get("_username",""))) / phone)
        client    = None

        pc_log(st, "info", f"Используем аккаунт [{phone[-4:]}]. Номеров для проверки: {len(phones)}")

        try:
            client = TelegramClient(sf, api_id, api_hash)
            await client.connect()
            if not await client.is_user_authorized():
                pc_log(st, "error", f"[{phone[-4:]}] Сессия не авторизована")
                st["running"] = False
                return

            # Проверяем пакетами по 10 номеров через ImportContacts
            batch_size = 10
            for i in range(0, len(phones), batch_size):
                if st["stop_flag"]: break

                batch = phones[i:i + batch_size]
                contacts_to_add = []
                for j, ph in enumerate(batch):
                    clean = ph.strip().replace(" ", "").replace("-", "")
                    if not clean.startswith("+"):
                        clean = "+" + clean
                    contacts_to_add.append(InputPhoneContact(
                        client_id=i + j,
                        phone=clean,
                        first_name="check",
                        last_name=""
                    ))

                try:
                    result = await client(ImportContactsRequest(contacts_to_add))
                    imported = {u.phone: u for u in result.users if hasattr(u, 'phone') and u.phone}

                    for ph_input, contact in zip(batch, contacts_to_add):
                        clean = contact.phone
                        st["cur"] = clean
                        st["progress"] += 1

                        # Ищем пользователя по номеру среди импортированных
                        matched = None
                        for uid, u in imported.items():
                            norm_uid = "+"+uid.lstrip("+")
                            norm_ph  = "+"+clean.lstrip("+")
                            if norm_uid == norm_ph or uid == clean.lstrip("+"):
                                matched = u
                                break

                        # Если не нашли по прямому совпадению — ищем среди всех users
                        if not matched:
                            for u in result.users:
                                uph = getattr(u, 'phone', '') or ''
                                if uph and (uph == clean.lstrip("+") or "+"+uph == clean):
                                    matched = u
                                    break

                        if matched:
                            st["found"] += 1
                            user_data = {
                                "phone":      clean,
                                "id":         matched.id,
                                "first_name": matched.first_name or "",
                                "last_name":  matched.last_name or "",
                                "username":   matched.username or "",
                                "is_premium": bool(getattr(matched, "premium", False)),
                                "has_avatar": bool(matched.photo),
                            }
                            st["results"].append(user_data)
                            name = (matched.first_name or "") + (" " + matched.last_name if matched.last_name else "")
                            uname = f" @{matched.username}" if matched.username else ""
                            pc_log(st, "success", f"✓ {clean} → {name.strip()}{uname} (id:{matched.id})")
                        else:
                            st["notfound"] += 1
                            pc_log(st, "info", f"✗ {clean} — не в Telegram")

                    # Удаляем временные контакты чтобы не засорять книгу
                    if result.users:
                        try:
                            await client(DeleteContactsRequest(id=[u.id for u in result.users]))
                        except Exception:
                            pass

                except errors.FloodWaitError as fe:
                    pc_log(st, "warn", f"FloodWait {fe.seconds}s — жду...")
                    await asyncio.sleep(min(fe.seconds, 60))
                    st["progress"] += len(batch)
                except Exception as ex:
                    pc_log(st, "error", f"Ошибка пакета {i}-{i+len(batch)}: {str(ex)[:100]}")
                    st["errors"] += len(batch)
                    st["progress"] += len(batch)

                await asyncio.sleep(delay)

        except Exception as ex:
            pc_log(st, "error", f"Критическая ошибка: {str(ex)[:120]}")
        finally:
            if client:
                try: await client.disconnect()
                except: pass

        st["running"]   = False
        st["stop_flag"] = False
        pc_log(st, "success",
               f"═══ Готово. Найдено: {st['found']} / {st['total']}, "
               f"не найдено: {st['notfound']}, ошибок: {st['errors']} ═══")

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(_run())
    finally:
        loop.close()


@app.route("/api/phonechecker/start", methods=["POST"])
@require_license
def api_phonechecker_start():
    from flask import g
    st = _upc(g.username)

    if st["running"]:
        st["running"] = False
        import time; time.sleep(0.2)

    data      = (request.get_json(force=True, silent=True) or {})
    phones    = data.get("phones", [])
    if not phones:
        return jsonify({"ok": False, "error": "Нет номеров"}), 400

    pc_state.update({
        "running": True, "stop_flag": False,
        "progress": 0, "total": len(phones),
        "found": 0, "notfound": 0, "errors": 0,
        "log": [], "results": [], "cur": "",
    })
    st["_username"] = g.username
    threading.Thread(
        target=phonechecker_task,
        kwargs=dict(
            phones=phones,
            acc_phones=data.get("acc_phones", []),
            delay=int(data.get("delay", 3)),
            st=st,
        ),
        daemon=True
    ).start()
    return jsonify({"ok": True})


@app.route("/api/phonechecker/stop", methods=["POST"])
@require_license
def api_phonechecker_stop():
    from flask import g
    st = _upc(g.username)

    st["stop_flag"] = True
    return jsonify({"ok": True})


@app.route("/api/phonechecker/status")
@require_license
def api_phonechecker_status():
    from flask import g
    st = _upc(g.username)

    return jsonify({
        "running":  st["running"],
        "progress": st["progress"],
        "total":    st["total"],
        "found":    st["found"],
        "notfound": st["notfound"],
        "errors":   st["errors"],
        "cur":      st["cur"],
        "log":      st["log"],
        "results":  st["results"],
    })


@app.route("/api/phonechecker/export", methods=["POST"])
@require_license
def api_phonechecker_export():
    from flask import g
    st = _upc(g.username)

    data    = (request.get_json(force=True, silent=True) or {})
    results = data.get("results", st["results"])
    if not results:
        return jsonify({"ok": False, "error": "Нет данных"}), 400
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"ok": False, "error": "pip install openpyxl"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phone Check"

    headers = ["Номер", "Имя", "Фамилия", "Username", "ID", "Premium", "Аватар"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1a1e27")
        cell.alignment = Alignment(horizontal="center")

    for row, u in enumerate(results, 2):
        ws.append([
            u.get("phone", ""),
            u.get("first_name", ""),
            u.get("last_name", ""),
            "@" + u["username"] if u.get("username") else "",
            u.get("id", ""),
            "Да" if u.get("is_premium") else "Нет",
            "Да" if u.get("has_avatar") else "Нет",
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    Path(PARSER_RESULTS_DIR).mkdir(exist_ok=True)
    fname    = f"phone_check_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fpath    = Path(PARSER_RESULTS_DIR) / fname
    wb.save(str(fpath))
    return jsonify({"ok": True, "url": f"/parser_results/{fname}", "filename": fname})


# ── Рассылка по контактам ─────────────────────────────────────────────────

bc_contacts_state = {
    "running": False, "stop_flag": False,
    "progress": 0, "total": 0,
    "sent": 0, "fail": 0, "skip": 0,
    "log": [], "cur": "",
}


def bcc_log(st, level, msg):
    _ulog(st["log"], level, msg, "BCC")


def bc_contacts_task(text, phones_filter, delay, limit_per_acc, st):
    async def _run():
        try:
            from telethon import TelegramClient, errors
            from telethon.tl.functions.contacts import GetContactsRequest
        except ImportError:
            bcc_log(st, "error", "Telethon не установлен")
            st["running"] = False
            return

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)
        accounts = get_accounts(st.get("_username",""))
        if phones_filter:
            accounts = [a for a in accounts if a["phone"] in phones_filter]
        if not accounts:
            bcc_log(st, "error", "Нет аккаунтов")
            st["running"] = False
            return

        bcc_log(st, "success", f"▶ Старт. Аккаунтов: {len(accounts)}")
        st["total"] = len(accounts)

        for acc_idx, acc in enumerate(accounts):
            if st["stop_flag"]:
                bcc_log(st, "warn", "⏹ Остановлено")
                break

            phone = acc["phone"]
            st["progress"] = acc_idx + 1
            client = None
            try:
                client = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / phone), api_id, api_hash)
                await client.connect()
                if not await client.is_user_authorized():
                    bcc_log(st, "warn", f"[{phone[-4:]}] Не авторизован, пропуск")
                    continue

                # Получаем контакты аккаунта
                contacts_result = await client(GetContactsRequest(hash=0))
                contacts = [u for u in contacts_result.users if not getattr(u, "bot", False)]

                if limit_per_acc and limit_per_acc > 0:
                    contacts = contacts[:limit_per_acc]

                bcc_log(st, "info", f"[{phone[-4:]}] Контактов: {len(contacts)}")

                sent_count = 0
                for user in contacts:
                    if st["stop_flag"]: break

                    st["cur"] = f"[{phone[-4:]}] → {user.username or user.id}"

                    try:
                        await client.send_message(user.id, text, parse_mode='html')
                        st["sent"] += 1
                        sent_count += 1
                        bcc_log(st, "success", f"[{phone[-4:]}] ✓ → {user.username or user.first_name or user.id}")
                        await asyncio.sleep(delay)

                    except errors.FloodWaitError as fe:
                        bcc_log(st, "warn", f"[{phone[-4:]}] FloodWait {fe.seconds}s")
                        await asyncio.sleep(min(fe.seconds, 60))
                    except errors.UserPrivacyRestrictedError:
                        bcc_log(st, "warn", f"[{phone[-4:]}] ✗ {user.username or user.id} — приватность")
                        st["fail"] += 1
                    except errors.InputUserDeactivatedError:
                        bcc_log(st, "warn", f"[{phone[-4:]}] ✗ {user.id} — аккаунт удалён")
                        st["fail"] += 1
                    except Exception as ex:
                        bcc_log(st, "error", f"[{phone[-4:]}] ✗ {user.id}: {str(ex)[:80]}")
                        st["fail"] += 1

                bcc_log(st, "success", f"[{phone[-4:]}] Завершено — отправлено: {sent_count}")

            except Exception as ex:
                bcc_log(st, "error", f"[{phone[-4:]}] Ошибка: {str(ex)[:100]}")
            finally:
                if client:
                    try: await client.disconnect()
                    except: pass

        st["running"]   = False
        st["stop_flag"] = False
        bcc_log(st, "success", f"═══ Рассылка по контактам завершена. Отправлено: {st['sent']}, ошибок: {st['fail']}, пропущено: {st['skip']} ═══")

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(_run())
    finally:
        loop.close()


@app.route("/api/broadcast/contacts/start", methods=["POST"])
@require_license
def api_bc_contacts_start():
    from flask import g
    st = _ubcc(g.username)

    if st["running"]:
        return jsonify({"ok": False, "error": "Рассылка уже запущена"}), 400
    data = (request.get_json(force=True, silent=True) or {})
    text = data.get("text", "").strip()
    if not text:
        return jsonify({"ok": False, "error": "Текст пуст"}), 400

    phones_filter = data.get("phones", [])
    accounts      = get_accounts(g.username)
    if phones_filter:
        accounts = [a for a in accounts if a["phone"] in phones_filter]
    if not accounts:
        return jsonify({"ok": False, "error": "Нет аккаунтов"}), 400

    bc_contacts_state.update({
        "running": True, "stop_flag": False,
        "progress": 0, "total": len(accounts),
        "sent": 0, "fail": 0, "skip": 0,
        "log": [], "cur": "",
    })

    st["_username"] = g.username
    threading.Thread(
        target=bc_contacts_task,
        kwargs=dict(
            text=text,
            phones_filter=phones_filter,
            delay=int(data.get("delay", 5)),
            limit_per_acc=int(data.get("limit_per_acc", 0)),
        ),
        daemon=True
    ).start()
    return jsonify({"ok": True, "accounts": len(accounts)})


@app.route("/api/broadcast/contacts/stop", methods=["POST"])
@require_license
def api_bc_contacts_stop():
    from flask import g
    st = _ubcc(g.username)

    st["stop_flag"] = True
    bcc_log(st, "warn", "Запрос остановки...")
    return jsonify({"ok": True})


@app.route("/api/broadcast/contacts/status")
@require_license
def api_bc_contacts_status():
    from flask import g
    st = _ubcc(g.username)

    return jsonify({
        "running":  st["running"],
        "progress": st["progress"],
        "total":    st["total"],
        "sent":     st["sent"],
        "fail":     st["fail"],
        "skip":     st["skip"],
        "cur":      st["cur"],
        "log":      st["log"],
    })


# ── Масслайкинг сторис ────────────────────────────────────────────────────

masslook_state = {
    "running": False, "stop_flag": False,
    "progress": 0, "total": 0,
    "viewed": 0, "liked": 0, "users_done": 0, "errors": 0,
    "log": [], "cur_acc": "", "cur_user": "",
}


def ml_log(st, level, msg):
    _ulog(st["log"], level, msg, "ML")


def masslook_task(users, phones, stories_per_user, view_delay, user_delay,
                  user_limit, reaction_mode, reaction_emoji, reaction_emojis, st=None):
    import random as _random

    async def _run():
        try:
            from telethon import TelegramClient, errors
            from telethon.tl.functions.stories import (
                GetPeerStoriesRequest, SendReactionRequest, ReadStoriesRequest
            )
            from telethon.tl.types import ReactionEmoji
        except ImportError:
            ml_log(st, "error", "Telethon не установлен или версия не поддерживает сторис")
            st["running"] = False
            return

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)
        accounts = get_accounts(st.get("_username","") if st else "")
        if phones:
            accounts = [a for a in accounts if a["phone"] in phones]
        if not accounts:
            ml_log(st, "error", "Нет аккаунтов")
            st["running"] = False
            return

        target_users = users
        if user_limit and user_limit > 0:
            target_users = users[:user_limit]

        st["total"]      = len(target_users)
        st["progress"]   = 0
        st["viewed"]     = 0
        st["liked"]      = 0
        st["users_done"] = 0
        st["errors"]     = 0

        ml_log(st, "success", f"▶ Старт. Пользователей: {len(target_users)}, аккаунтов: {len(accounts)}, реакция: {reaction_mode}")

        acc_idx = 0

        for idx, user_ref in enumerate(target_users):
            if st["stop_flag"]:
                ml_log(st, "warn", "⏹ Остановлено")
                break

            phone = accounts[acc_idx % len(accounts)]["phone"]
            acc_idx = (acc_idx + 1) % len(accounts)

            st["cur_acc"]  = phone[-4:]
            st["cur_user"] = user_ref
            st["progress"] = idx + 1

            client = None
            try:
                client = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / phone), api_id, api_hash)
                await client.connect()
                if not await client.is_user_authorized():
                    ml_log(st, "warn", f"[{phone[-4:]}] Не авторизован, пропуск")
                    st["errors"] += 1
                    continue

                # Получаем entity пользователя
                try:
                    if user_ref.lstrip("-").isdigit():
                        peer = await client.get_entity(int(user_ref))
                    else:
                        peer = await client.get_entity(user_ref.lstrip("@"))
                except Exception as ue:
                    ml_log(st, "warn", f"Не найден {user_ref}: {str(ue)[:60]}")
                    st["errors"] += 1
                    st["users_done"] += 1
                    continue

                # Получаем сторис пользователя
                try:
                    result = await client(GetPeerStoriesRequest(peer=peer))
                    stories = result.stories.stories if result.stories else []
                except Exception as se:
                    ml_log(st, "warn", f"[{phone[-4:]}] Нет сторис у {user_ref}: {str(se)[:60]}")
                    st["users_done"] += 1
                    continue

                if not stories:
                    ml_log(st, "info", f"[{phone[-4:]}] {user_ref} — сторис нет")
                    st["users_done"] += 1
                    continue

                # Ограничиваем кол-во сторис
                if stories_per_user and stories_per_user > 0:
                    stories = stories[:stories_per_user]

                ml_log(st, "info", f"[{phone[-4:]}] {user_ref} — {len(stories)} сторис")

                story_ids = [s.id for s in stories]

                # Читаем (просматриваем) сторис
                try:
                    await client(ReadStoriesRequest(peer=peer, max_id=max(story_ids)))
                    st["viewed"] += len(stories)
                    ml_log(st, "success", f"[{phone[-4:]}] Просмотрено {len(stories)} сторис у {user_ref}")
                except Exception as re:
                    ml_log(st, "warn", f"[{phone[-4:]}] Ошибка просмотра: {str(re)[:60]}")

                await asyncio.sleep(view_delay)

                # Ставим реакцию если нужно
                if reaction_mode != "none":
                    for story in stories:
                        if st["stop_flag"]: break
                        try:
                            if reaction_mode == "like":
                                emoji = "❤"
                            elif reaction_mode == "random":
                                emoji_raw = _random.choice(reaction_emojis) if reaction_emojis else "❤️"
                                # Убираем вариационный селектор для Telegram
                                emoji = emoji_raw.replace("\ufe0f","")
                            elif reaction_mode == "custom":
                                emoji = reaction_emoji.replace("\ufe0f","")
                            else:
                                emoji = "❤"

                            await client(SendReactionRequest(
                                peer=peer,
                                story_id=story.id,
                                reaction=ReactionEmoji(emoticon=emoji)
                            ))
                            st["liked"] += 1
                            await asyncio.sleep(view_delay)
                        except errors.FloodWaitError as fe:
                            ml_log(st, "warn", f"[{phone[-4:]}] FloodWait {fe.seconds}s")
                            await asyncio.sleep(min(fe.seconds, 30))
                        except Exception as ex:
                            ml_log(st, "warn", f"[{phone[-4:]}] Реакция на сторис {story.id}: {str(ex)[:60]}")

                st["users_done"] += 1

            except errors.FloodWaitError as fe:
                ml_log(st, "warn", f"[{phone[-4:]}] FloodWait {fe.seconds}s")
                st["errors"] += 1
                await asyncio.sleep(min(fe.seconds, 60))
            except Exception as ex:
                ml_log(st, "error", f"[{phone[-4:]}] {user_ref}: {str(ex)[:100]}")
                st["errors"] += 1
            finally:
                if client:
                    try: await client.disconnect()
                    except: pass

            await asyncio.sleep(user_delay)

        st["running"]   = False
        st["stop_flag"] = False
        ml_log(st, "success", f"═══ Завершено. Просмотрено: {st['viewed']}, лайков: {st['liked']}, пользователей: {st['users_done']} ═══")

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(_run())
    finally:
        loop.close()


@app.route("/api/masslook/start", methods=["POST"])
@require_license
def api_masslook_start():
    from flask import g
    st = _uml(g.username)

    if st["running"]:
        return jsonify({"ok": False, "error": "Уже запущен"}), 400
    data  = (request.get_json(force=True, silent=True) or {})
    users = data.get("users", [])
    if not users:
        return jsonify({"ok": False, "error": "Нет пользователей"}), 400

    st["_username"] = g.username
    st.update({
        "running": True, "stop_flag": False,
        "progress": 0, "total": len(users),
        "viewed": 0, "liked": 0, "users_done": 0, "errors": 0,
        "log": [], "cur_acc": "", "cur_user": "",
    })

    threading.Thread(
        target=masslook_task,
        kwargs=dict(
            st=st,
            users=users,
            phones=data.get("phones", []),
            stories_per_user=int(data.get("stories_per_user", 0)),
            view_delay=int(data.get("view_delay", 3)),
            user_delay=int(data.get("user_delay", 5)),
            user_limit=int(data.get("user_limit", 0)),
            reaction_mode=data.get("reaction_mode", "none"),
            reaction_emoji=data.get("reaction_emoji", "❤"),
            reaction_emojis=data.get("reaction_emojis", ["❤️"]),
        ),
        daemon=True
    ).start()
    return jsonify({"ok": True})


@app.route("/api/masslook/stop", methods=["POST"])
@require_license
def api_masslook_stop():
    from flask import g
    st = _uml(g.username)

    st["stop_flag"] = True
    ml_log(st, "warn", "Запрос остановки...")
    return jsonify({"ok": True})


@app.route("/api/masslook/status")
@require_license
def api_masslook_status():
    from flask import g
    st = _uml(g.username)

    return jsonify({
        "running":    st["running"],
        "progress":   st["progress"],
        "total":      st["total"],
        "viewed":     st["viewed"],
        "liked":      st["liked"],
        "users_done": st["users_done"],
        "errors":     st["errors"],
        "cur_acc":    st["cur_acc"],
        "cur_user":   st["cur_user"],
        "log":        st["log"],
    })


# ── Базы парсинга ─────────────────────────────────────────────────────────

@app.route("/api/bases/save", methods=["POST"])
@require_license
def api_bases_save():
    data    = (request.get_json(force=True, silent=True) or {})
    name    = data.get("name", "").strip()
    results = data.get("results", [])
    if not name:
        return jsonify({"ok": False, "error": "Нет названия"})
    if not results:
        return jsonify({"ok": False, "error": "Нет данных"})
    # Убираем служебные поля
    clean = [{k:v for k,v in u.items() if k != "_last_online_dt"} for u in results]
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    bid = ts + "_" + re.sub(r'[^\w]', '_', name)[:40]
    meta = {
        "id":    bid,
        "name":  name,
        "count": len(clean),
        "date":  datetime.now().strftime("%d.%m.%Y %H:%M"),
        "ts":    ts,
    }
    base_dir = Path(BASES_DIR) / bid
    base_dir.mkdir(exist_ok=True)
    (base_dir / "meta.json").write_text(json.dumps(meta, ensure_ascii=False))
    (base_dir / "data.json").write_text(json.dumps(clean, ensure_ascii=False))
    log("success", f"База сохранена: «{name}» ({len(clean)} чел.)")
    return jsonify({"ok": True, "id": bid})


@app.route("/api/bases/list")
@require_license
def api_bases_list():
    bases = []
    for d in sorted(Path(BASES_DIR).iterdir(), reverse=True):
        meta_f = d / "meta.json"
        if meta_f.exists():
            try:
                bases.append(json.loads(meta_f.read_text(encoding="utf-8")))
            except Exception:
                pass
    return jsonify({"bases": bases})


@app.route("/api/bases/get/<base_id>")
@require_license
def api_bases_get(base_id):
    base_dir = Path(BASES_DIR) / base_id
    meta_f   = base_dir / "meta.json"
    data_f   = base_dir / "data.json"
    if not base_dir.exists() or not data_f.exists():
        return jsonify({"ok": False, "error": "База не найдена"})
    try:
        meta    = json.loads(meta_f.read_text(encoding="utf-8"))
        results = json.loads(data_f.read_text(encoding="utf-8"))
        return jsonify({"ok": True, "name": meta["name"], "results": results})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)})


@app.route("/api/bases/delete", methods=["POST"])
@require_license
def api_bases_delete():
    bid      = ((request.get_json(force=True, silent=True) or {})).get("id", "").strip()
    base_dir = Path(BASES_DIR) / bid
    if not base_dir.exists():
        return jsonify({"ok": False, "error": "База не найдена"})
    try:
        import shutil
        shutil.rmtree(str(base_dir))
        log("warn", f"База удалена: {bid}")
        return jsonify({"ok": True})
    except Exception as ex:
        return jsonify({"ok": False, "error": str(ex)})


# ── Настройки ──────────────────────────────────────────────────────────────

@app.route("/api/settings", methods=["GET"])
@require_license
def api_settings_get():
    return jsonify(settings)


@app.route("/api/settings", methods=["POST"])
@require_license
def api_settings_save():
    data = (request.get_json(force=True, silent=True) or {})
    settings.update(data)
    state["admin"] = data.get("admin", state["admin"])
    state["delay"] = int(data.get("delay", state["delay"]))
    # Per-user state sync
    from flask import g
    _ust(g.username)["admin"] = state["admin"]
    _ust(g.username)["delay"] = state["delay"]
    # Нормализуем smtp_port и smtp_tls
    if "smtp_port" in data:
        try:
            settings["smtp_port"] = int(data["smtp_port"])
        except (ValueError, TypeError):
            settings["smtp_port"] = 587
    if "smtp_tls" in data:
        settings["smtp_tls"] = bool(data["smtp_tls"])
    Path(SETTINGS_FILE).write_text(json.dumps(settings, ensure_ascii=False, indent=2))
    log("success", "Настройки сохранены")
    return jsonify({"ok": True})


@app.route("/api/settings/smtp-test", methods=["POST"])
def api_settings_smtp_test():
    token = request.headers.get("X-Token", "")
    sessions = load_web_sessions()
    if token not in sessions:
        return jsonify({"ok": False, "error": "Не авторизован"}), 403
    data = request.get_json(force=True, silent=True) or {}
    to_addr = (data.get("email") or "").strip()
    if not to_addr:
        return jsonify({"ok": False, "error": "Укажите адрес получателя"})
    body = """
    <div style="font-family:sans-serif;padding:20px;background:#0c0e12;color:#e8eaf0;border-radius:12px;">
      <div style="font-size:20px;font-weight:700;color:#4ade80;margin-bottom:8px;">TG Channel Creator</div>
      <p>Тестовое письмо — SMTP настроен корректно ✓</p>
    </div>
    """
    ok, err = send_email(to_addr, "SMTP Test — TG Channel Creator", body)
    if ok:
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": err})


# ── Нейрокомментинг ────────────────────────────────────────────────────────

def neuro_log(st, level, msg):
    _ulog(st["neuro_log"], level, msg, "NEURO")


DEFAULT_DEEPSEEK_KEY = "sk-3ddc1fd76de24ca7939fb99cd009ad0d"
_neuro_last_ids: dict = {}   # channel -> last processed msg_id (глобальный)


def deepseek_generate(api_key, prompt, post_text, max_tokens=150):
    import urllib.request, json as _j, ssl
    payload = _j.dumps({
        "model": "deepseek-chat",
        "messages": [
            {"role": "system", "content": prompt},
            {"role": "user",   "content": f"Пост:\n{post_text[:1000]}"},
        ],
        "max_tokens": max_tokens,
        "temperature": 0.9,
    }).encode()
    req = urllib.request.Request(
        "https://api.deepseek.com/v1/chat/completions",
        data=payload,
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"},
        method="POST",
    )
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode    = ssl.CERT_NONE
    with urllib.request.urlopen(req, timeout=30, context=ctx) as r:
        return _j.loads(r.read())["choices"][0]["message"]["content"].strip()


@app.route("/api/neuro/check_key", methods=["POST"])
@require_license
def api_neuro_check_key():
    key = ((request.get_json(force=True, silent=True) or {})).get("api_key", "").strip() or DEFAULT_DEEPSEEK_KEY
    try:
        text = deepseek_generate(key, "Reply with one word: OK", "test", max_tokens=5)
        return jsonify({"ok": True, "valid": True, "model": "deepseek-chat", "response": text[:30]})
    except Exception as ex:
        return jsonify({"ok": False, "valid": False, "error": str(ex)[:120]})


@app.route("/api/neuro/log")
@require_license
def api_neuro_log():
    from flask import g
    st = _ust(g.username)

    since = int(request.args.get("since", 0))
    return jsonify({
        "log":    st["neuro_log"][since:],
        "total":  len(st["neuro_log"]),
        "active": st["neuro_active"],
    })


@app.route("/api/neuro/log/clear", methods=["POST"])
@require_license
def api_neuro_log_clear():
    from flask import g
    st = _ust(g.username)

    st["neuro_log"] = []
    return jsonify({"ok": True})


def neuro_worker(api_key, prompt, channels, monitors, commenters, delay_min, delay_max, interval, st=None):
    """Единственный воркер нейрокомментинга."""
    global _neuro_last_ids
    import random as _rnd

    async def _run():
        global _neuro_last_ids
        try:
            from telethon import TelegramClient, errors
        except ImportError:
            neuro_log(st, "error", "Telethon не установлен")
            st["neuro_active"] = False
            return

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)

        all_phones = {a["phone"] for a in get_accounts(st.get("_username","") if st else "")}
        mon_list = [p for p in monitors   if p in all_phones]
        com_list = [p for p in commenters if p in all_phones]

        if not mon_list:
            neuro_log(st, "error", f"Мониторы не найдены в sessions/. Телефоны: {monitors}")
            st["neuro_active"] = False
            return
        if not com_list:
            neuro_log(st, "error", f"Комментаторы не найдены в sessions/. Телефоны: {commenters}")
            st["neuro_active"] = False
            return

        neuro_log(st, "info", f"Мониторов: {len(mon_list)}, комментаторов: {len(com_list)}, каналов: {len(channels)}")

        # ── Инициализация: запоминаем текущий последний пост ──────────────────
        init_client = None
        try:
            init_client = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / mon_list[0]), api_id, api_hash)
            await init_client.connect()
            if await init_client.is_user_authorized():
                for ch in channels:
                    if ch in _neuro_last_ids:
                        neuro_log(st, "info", f"{ch}: продолжаю с ID={_neuro_last_ids[ch]}")
                        continue
                    try:
                        ent  = await init_client.get_entity(ch.lstrip("@"))
                        msgs = await init_client.get_messages(ent, limit=1)
                        _neuro_last_ids[ch] = msgs[0].id if msgs else 0
                        neuro_log(st, "info", f"{ch}: последний ID={_neuro_last_ids[ch]}, жду новых постов...")
                    except Exception as ex:
                        neuro_log(st, "warn", f"Инит {ch}: {ex}")
                        _neuro_last_ids[ch] = 0
        except Exception as ex:
            neuro_log(st, "warn", f"Ошибка инициализации: {ex}")
        finally:
            if init_client:
                try: await init_client.disconnect()
                except: pass

        neuro_log(st, "success", "✓ Готов. Слежу за каналами...")

        mon_idx = 0
        com_idx = 0

        while st["neuro_active"]:
            for ch in channels:
                if not st["neuro_active"]: break

                # 1. Монитор читает новые посты
                mon_phone  = mon_list[mon_idx % len(mon_list)]
                mon_idx   += 1
                mon_client = None
                new_posts  = []

                try:
                    mon_client = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / mon_phone), api_id, api_hash)
                    await mon_client.connect()
                    if not await mon_client.is_user_authorized():
                        neuro_log(st, "warn", f"[{mon_phone[-4:]}] Не авторизован")
                    else:
                        ent  = await mon_client.get_entity(ch.lstrip("@"))
                        msgs = await mon_client.get_messages(ent, limit=10)
                        last = _neuro_last_ids.get(ch, 0)
                        new_posts = [m for m in reversed(msgs)
                                     if m.text and len(m.text.strip()) >= 5 and m.id > last]
                        if new_posts:
                            neuro_log(st, "info", f"[{mon_phone[-4:]}] {ch}: {len(new_posts)} новых")
                        else:
                            neuro_log(st, "info", f"[{mon_phone[-4:]}] {ch}: нет новых (last={last})")
                except Exception as ex:
                    neuro_log(st, "error", f"Монитор {ch}: {str(ex)[:80]}")
                finally:
                    if mon_client:
                        try: await mon_client.disconnect()
                        except: pass

                # 2. Комментатор пишет под каждый новый пост
                for post in new_posts:
                    if not st["neuro_active"]: break

                    # Генерация текста
                    try:
                        comment = deepseek_generate(api_key, prompt, post.text)
                    except Exception as ai_ex:
                        neuro_log(st, "warn", f"DeepSeek: {str(ai_ex)[:80]}")
                        _neuro_last_ids[ch] = post.id
                        continue

                    # Отправка
                    com_phone  = com_list[com_idx % len(com_list)]
                    com_idx   += 1
                    com_client = None
                    try:
                        com_client = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / com_phone), api_id, api_hash)
                        await com_client.connect()
                        if not await com_client.is_user_authorized():
                            neuro_log(st, "warn", f"[{com_phone[-4:]}] Не авторизован")
                            _neuro_last_ids[ch] = post.id
                            continue
                        ent = await com_client.get_entity(ch.lstrip("@"))
                        sent = await com_client.send_message(ent, comment, comment_to=post.id)
                        _neuro_last_ids[ch] = post.id
                        chan_clean = ch.lstrip("@")
                        comment_url = f"https://t.me/{chan_clean}/{sent.id}"
                        neuro_log(st, "success", f"[{com_phone[-4:]}] ✓ {ch}: {comment[:60]}... | LINK:{comment_url}")
                        await asyncio.sleep(_rnd.randint(delay_min, max(delay_min, delay_max)))
                    except errors.ChatWriteForbiddenError:
                        neuro_log(st, "warn", f"{ch}: нет прав на комментирование")
                        _neuro_last_ids[ch] = post.id
                    except errors.FloodWaitError as fe:
                        neuro_log(st, "warn", f"[{com_phone[-4:]}] FloodWait {fe.seconds}s")
                        await asyncio.sleep(min(fe.seconds, 60))
                    except Exception as ex:
                        neuro_log(st, "error", f"[{com_phone[-4:]}] {str(ex)[:80]}")
                        _neuro_last_ids[ch] = post.id
                    finally:
                        if com_client:
                            try: await com_client.disconnect()
                            except: pass

            if st["neuro_active"]:
                await asyncio.sleep(interval)

        st["neuro_active"] = False
        neuro_log(st, "warn", "⏹ Остановлен")

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(_run())
    finally:
        loop.close()


@app.route("/api/neuro/start", methods=["POST"])
@require_license
def api_neuro_start():
    from flask import g
    st = _ust(g.username)

    global _neuro_last_ids
    # Сбрасываем если уже был запущен
    if st["neuro_active"]:
        st["neuro_active"] = False
        _neuro_last_ids = {}
        import time; time.sleep(0.5)

    data       = (request.get_json(force=True, silent=True) or {})
    api_key    = data.get("api_key", "").strip() or settings.get("deepseek_api_key") or DEFAULT_DEEPSEEK_KEY
    channels   = data.get("channels",   [])
    monitors   = data.get("monitors",   [])
    commenters = data.get("commenters", [])

    if not channels:   return jsonify({"ok": False, "error": "Нет каналов"}),      400
    if not monitors:   return jsonify({"ok": False, "error": "Нет мониторов"}),    400
    if not commenters: return jsonify({"ok": False, "error": "Нет комментаторов"}),400

    prompt = data.get("prompt", "Напиши короткий живой комментарий к посту от лица обычного пользователя на русском языке (1-2 предложения). Без шаблонных фраз.")
    d_min  = int(data.get("delay_min",  10))
    d_max  = int(data.get("delay_max",  60))
    intvl  = int(data.get("interval",   60))

    st["neuro_active"] = True
    neuro_log(st, "success", f"▶ Мониторов: {len(monitors)}, комментаторов: {len(commenters)}, каналов: {len(channels)}")

    st["_username"] = g.username
    threading.Thread(
        target=neuro_worker,
        args=(api_key, prompt, channels, monitors, commenters, d_min, d_max, intvl, st),
        daemon=True,
    ).start()
    return jsonify({"ok": True})


@app.route("/api/neuro/stop", methods=["POST"])
@require_license
def api_neuro_stop():
    from flask import g
    st = _ust(g.username)

    global _neuro_last_ids
    st["neuro_active"] = False
    _neuro_last_ids = {}
    neuro_log(st, "warn", "⏹ Остановлен")
    return jsonify({"ok": True})


# ── Авто-реакции ───────────────────────────────────────────────────────────

reactions_state = {
    "active": False,
    "stop_flag": False,
    "sent": 0,
    "errors": 0,
    "joined": 0,
    "log": [],
}
_reactions_last_ids: dict = {}   # ch -> last processed post_id (для режима new_posts)
_reactions_done: set = set()     # (ch, key) — уже обработанные посты/комментарии


def reactions_log(st, level, msg):
    _ulog(st["log"], level, msg, "REACT")


def reactions_worker(channels, monitors, workers, emojis, delay_min, delay_max,
                     interval, random_emoji, join_before,
                     comments_limit, watch_mode, posts_count, st, last_ids, done_set):
    """
    watch_mode: "new_posts"    — ждать новых постов (по last_id)
                "recent_posts" — каждый цикл брать последние posts_count постов
    Реакции ставятся только на комментарии (не на сами посты).
    """
    import random as _rnd

    async def _run():
        # last_ids and done_set are per-user, passed as parameters

        try:
            from telethon import TelegramClient, errors
            from telethon.tl.functions.messages import SendReactionRequest, GetRepliesRequest
            from telethon.tl.functions.channels import JoinChannelRequest, GetFullChannelRequest
            from telethon.tl.types import ReactionEmoji, Channel as TLChannel
        except ImportError:
            reactions_log(st, "error", "Telethon не установлен")
            st["active"] = False
            return

        api_id   = int(settings.get("api_id", API_ID))
        api_hash = settings.get("api_hash", API_HASH)

        all_phones = {a["phone"] for a in get_accounts(st.get("_username",""))}
        mon_list = [p for p in monitors if p in all_phones]
        wrk_list = [p for p in workers  if p in all_phones]

        if not mon_list:
            reactions_log(st, "error", "Нет аккаунтов-мониторов в sessions/")
            st["active"] = False
            return
        if not wrk_list:
            reactions_log(st, "error", "Нет рабочих аккаунтов в sessions/")
            st["active"] = False
            return

        watch_lbl = {"new_posts": "новые посты", "recent_posts": f"последние {posts_count} постов"}.get(watch_mode, watch_mode)
        reactions_log(st, "info", f"Мониторов: {len(mon_list)}, рабочих: {len(wrk_list)}, чатов: {len(channels)}")
        reactions_log(st, "info", f"Реакции на комментарии | Мониторинг: {watch_lbl} | Вступление: {'да' if join_before else 'нет'}")

        # ── Помощники ──────────────────────────────────────────────────────

        def pick_emoji():
            raw = _rnd.choice(emojis) if random_emoji and emojis else (emojis[0] if emojis else "❤️")
            return raw.replace("\ufe0f", "")

        async def do_join(client, ent, label):
            """Вступить в чат (игнорируем если уже вступили)."""
            try:
                await client(JoinChannelRequest(ent))
                st["joined"] += 1
                reactions_log(st, "info", f"Вступил в {label}")
            except Exception:
                pass

        async def react_to_msg(client, peer, msg_id, label):
            """Поставить реакцию на сообщение, вернуть True при успехе."""
            try:
                await client(SendReactionRequest(
                    peer=peer,
                    msg_id=msg_id,
                    reaction=[ReactionEmoji(emoticon=pick_emoji())]
                ))
                st["sent"] += 1
                reactions_log(st, "success", f"[wrk] {pick_emoji()} → {label}")
                await asyncio.sleep(_rnd.randint(delay_min, max(delay_min, delay_max)))
                return True
            except errors.ReactionInvalidError:
                reactions_log(st, "warn", f"[wrk] Реакция недоступна: {label}")
            except errors.FloodWaitError as fe:
                reactions_log(st, "warn", f"[wrk] FloodWait {fe.seconds}s")
                await asyncio.sleep(min(fe.seconds, 60))
            except Exception as ex:
                reactions_log(st, "warn", f"[wrk] {label}: {str(ex)[:70]}")
                st["errors"] += 1
            return False

        async def get_comments(client, ch_ent, post_id):
            """
            Возвращает (peer_for_reaction, comments_list).
            peer_for_reaction — куда слать SendReactionRequest.
            Для каналов (broadcast) — linked discussion group.
            Для групп (megagroup/chat) — сама группа.
            """
            is_broadcast = isinstance(ch_ent, TLChannel) and getattr(ch_ent, 'broadcast', False)
            if is_broadcast:
                # Канал → ищем linked discussion group
                try:
                    full = await client(GetFullChannelRequest(channel=ch_ent))
                    disc_id = getattr(full.full_chat, "linked_chat_id", None)
                    if not disc_id:
                        return None, []
                    disc_ent = await client.get_entity(disc_id)
                    if join_before:
                        await do_join(client, disc_ent, "группа обсуждений")
                    replies = await client(GetRepliesRequest(
                        peer=ch_ent, msg_id=post_id,
                        offset_id=0, offset_date=None,
                        add_offset=0, limit=comments_limit,
                        max_id=0, min_id=0, hash=0
                    ))
                    return disc_ent, (replies.messages if hasattr(replies, "messages") else [])
                except Exception as ex:
                    reactions_log(st, "warn", f"Комментарии канала: {str(ex)[:70]}")
                    return None, []
            else:
                # Группа/супергруппа → реплаи внутри той же группы
                try:
                    replies = await client(GetRepliesRequest(
                        peer=ch_ent, msg_id=post_id,
                        offset_id=0, offset_date=None,
                        add_offset=0, limit=comments_limit,
                        max_id=0, min_id=0, hash=0
                    ))
                    return ch_ent, (replies.messages if hasattr(replies, "messages") else [])
                except Exception as ex:
                    reactions_log(st, "warn", f"Комментарии группы: {str(ex)[:70]}")
                    return None, []

        # ── Инициализация last_ids для режима new_posts ────────────────────
        if watch_mode == "new_posts":
            init_client = None
            try:
                init_client = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / mon_list[0]), api_id, api_hash)
                await init_client.connect()
                if await init_client.is_user_authorized():
                    for ch in channels:
                        if ch in _reactions_last_ids:
                            continue
                        try:
                            ent  = await init_client.get_entity(ch.lstrip("@"))
                            msgs = await init_client.get_messages(ent, limit=1)
                            last_ids[ch] = msgs[0].id if msgs else 0
                            reactions_log(st, "info", f"{ch}: последний ID={last_ids[ch]}, жду новых постов...")
                        except Exception as ex:
                            reactions_log(st, "warn", f"Инит {ch}: {str(ex)[:70]}")
                            last_ids[ch] = 0
            except Exception as ex:
                reactions_log(st, "warn", f"Инициализация: {str(ex)[:70]}")
            finally:
                if init_client:
                    try: await init_client.disconnect()
                    except: pass

        reactions_log(st, "success", "✓ Слежу за чатами...")

        mon_idx = 0
        wrk_idx = 0

        while st["active"] and not st["stop_flag"]:
            for ch in channels:
                if not st["active"] or st["stop_flag"]:
                    break

                # ── Монитор получает посты ─────────────────────────────────
                mon_phone  = mon_list[mon_idx % len(mon_list)]
                mon_idx   += 1
                mon_client = None
                target_posts = []
                try:
                    mon_client = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / mon_phone), api_id, api_hash)
                    await mon_client.connect()
                    if not await mon_client.is_user_authorized():
                        reactions_log(st, "warn", f"[mon {mon_phone[-4:]}] Не авторизован")
                    else:
                        ent  = await mon_client.get_entity(ch.lstrip("@"))
                        if watch_mode == "new_posts":
                            fetch_limit = 10
                            msgs = await mon_client.get_messages(ent, limit=fetch_limit)
                            last = _reactions_last_ids.get(ch, 0)
                            target_posts = [m for m in reversed(msgs) if m.id > last]
                            status = f"{len(target_posts)} новых" if target_posts else f"нет новых (last={last})"
                        else:  # recent_posts
                            msgs = await mon_client.get_messages(ent, limit=posts_count)
                            target_posts = list(reversed(msgs))
                            status = f"взял {len(target_posts)} последних постов"
                        reactions_log(st, "info", f"[mon {mon_phone[-4:]}] {ch}: {status}")
                except Exception as ex:
                    reactions_log(st, "error", f"Монитор {ch}: {str(ex)[:80]}")
                finally:
                    if mon_client:
                        try: await mon_client.disconnect()
                        except: pass

                # ── Рабочий обрабатывает каждый пост ──────────────────────
                for post in target_posts:
                    if not st["active"] or st["stop_flag"]:
                        break

                    post_key = (ch, f"p{post.id}")
                    wrk_phone  = wrk_list[wrk_idx % len(wrk_list)]
                    wrk_idx   += 1
                    wrk_client = None
                    try:
                        wrk_client = TelegramClient(str(Path(get_sessions_dir(st.get("_username",""))) / wrk_phone), api_id, api_hash)
                        await wrk_client.connect()
                        if not await wrk_client.is_user_authorized():
                            reactions_log(st, "warn", f"[wrk {wrk_phone[-4:]}] Не авторизован")
                            if watch_mode == "new_posts":
                                last_ids[ch] = post.id
                            continue

                        ch_ent = await wrk_client.get_entity(ch.lstrip("@"))

                        # Вступаем в канал/группу
                        if join_before:
                            await do_join(wrk_client, ch_ent, ch)

                        # ── Реакции на комментарии ─────────────────────────
                        react_peer, comments = await get_comments(wrk_client, ch_ent, post.id)
                        if react_peer and comments:
                            new_comments = [c for c in comments
                                            if (ch, f"c{c.id}") not in done_set]
                            if new_comments:
                                reactions_log(st, "info", f"[wrk {wrk_phone[-4:]}] {ch}/{post.id}: "
                                                      f"{len(new_comments)} новых комментариев")
                            for comment in new_comments:
                                if not st["active"] or st["stop_flag"]:
                                    break
                                ckey = (ch, f"c{comment.id}")
                                ok = await react_to_msg(
                                    wrk_client, react_peer, comment.id,
                                    f"комментарий {ch}/{post.id}/#{comment.id}"
                                )
                                if ok:
                                    done_set.add(ckey)

                        if watch_mode == "new_posts":
                            last_ids[ch] = post.id

                        # Ограничиваем размер done_set
                        if len(done_set) > 5000:
                            done_set = set(list(done_set)[-3000:])

                    except errors.FloodWaitError as fe:
                        reactions_log(st, "warn", f"[wrk {wrk_phone[-4:]}] FloodWait {fe.seconds}s")
                        await asyncio.sleep(min(fe.seconds, 60))
                    except Exception as ex:
                        reactions_log(st, "error", f"[wrk {wrk_phone[-4:]}] {str(ex)[:80]}")
                        st["errors"] += 1
                        if watch_mode == "new_posts":
                            last_ids[ch] = post.id
                    finally:
                        if wrk_client:
                            try: await wrk_client.disconnect()
                            except: pass

            if st["active"] and not st["stop_flag"]:
                await asyncio.sleep(interval)

        st["active"] = False
        reactions_log(st, "warn", "⏹ Остановлен")

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(_run())
    finally:
        loop.close()


@app.route("/api/reactions/start", methods=["POST"])
@require_license
def api_reactions_start():
    from flask import g
    st = _ureact(g.username)

    global _reactions_last_ids, _reactions_done
    if st["active"]:
        st["active"] = False
        st["stop_flag"] = True
        time.sleep(0.5)

    data           = request.get_json(force=True, silent=True) or {}
    channels       = data.get("channels",  [])
    monitors       = data.get("monitors",  [])
    workers        = data.get("workers",   [])
    emojis         = data.get("emojis",    ["❤️"])
    delay_min      = int(data.get("delay_min",      5))
    delay_max      = int(data.get("delay_max",      30))
    interval       = int(data.get("interval",       60))
    random_emoji   = bool(data.get("random_emoji",  True))
    join_before    = bool(data.get("join_before",   False))
    comments_limit = int(data.get("comments_limit", 10))
    watch_mode     = data.get("watch_mode",   "new_posts")   # new_posts | recent_posts
    posts_count    = int(data.get("posts_count",    5))

    if not channels: return jsonify({"ok": False, "error": "Нет каналов/групп"}), 400
    if not monitors: return jsonify({"ok": False, "error": "Нет аккаунтов-мониторов"}), 400
    if not workers:  return jsonify({"ok": False, "error": "Нет рабочих аккаунтов"}), 400
    if not emojis:   return jsonify({"ok": False, "error": "Выберите хотя бы одну реакцию"}), 400

    reactions_state.update({"active": True, "stop_flag": False, "sent": 0, "errors": 0, "joined": 0, "log": []})
    _reactions_last_ids = {}
    _reactions_done = set()
    reactions_log(st, "success", f"▶ Старт. Чатов: {len(channels)}, мониторов: {len(monitors)}, рабочих: {len(workers)}")

    st["_username"] = g.username
    threading.Thread(
        target=reactions_worker,
        args=(channels, monitors, workers, emojis, delay_min, delay_max,
              interval, random_emoji, join_before,
              comments_limit, watch_mode, posts_count, st, _u_reactions_last_ids[g.username], _u_reactions_done[g.username]),
        daemon=True,
    ).start()
    return jsonify({"ok": True})


@app.route("/api/reactions/stop", methods=["POST"])
@require_license
def api_reactions_stop():
    from flask import g
    st = _ureact(g.username)

    global _reactions_last_ids
    st["active"]    = False
    st["stop_flag"] = True
    _reactions_last_ids = {}
    reactions_log(st, "warn", "⏹ Остановлен")
    return jsonify({"ok": True})


@app.route("/api/reactions/log")
@require_license
def api_reactions_log():
    from flask import g
    st = _ureact(g.username)

    offset = int(request.args.get("offset", 0))
    logs   = st["log"][offset:]
    return jsonify({
        "log":    logs,
        "offset": offset + len(logs),
        "active":  st["active"],
        "sent":    st["sent"],
        "errors":  st["errors"],
        "joined":  st["joined"],
    })


# ── Статистика ─────────────────────────────────────────────────────────────

@app.route("/api/stats")
@require_license
def api_stats():
    from flask import g
    accounts = get_accounts(g.username)
    valid    = sum(1 for a in accounts if a["status"] == "valid")
    total_ch = sum(a["channels"] for a in accounts)
    from flask import g
    st = _ust(g.username)
    results_ok = sum(1 for r in st["results"] if r.get("success"))
    return jsonify({
        "accounts_total":   len(accounts),
        "accounts_active":  valid if valid else len(accounts),
        "channels_created": total_ch,
        "queue_count":      len(st["queue"]),
        "results_count":    len(st["results"]),
        "results_ok":       results_ok,
        "running":          st["running"],
        "admin":            _ust(g.username)["admin"],
        "delay":            _ust(g.username)["delay"],
        "broadcast_runs":   _ust(g.username)["broadcast_runs"],
        "broadcast_ok":     _ust(g.username)["broadcast_ok"],
        "broadcast_fail":   _ust(g.username)["broadcast_fail"],
        "neuro_active":     _ust(g.username)["neuro_active"],
        "reactions_active": _ureact(g.username)["active"],
        "reactions_sent":   _ureact(g.username)["sent"],
    })


# ── Авторизация (лицензионные ключи) ──────────────────────────────────────

@app.route("/api/auth/login", methods=["POST"])
def api_auth_login():
    data         = request.get_json(force=True, silent=True) or {}
    key          = (data.get("key") or "").strip()
    device_token = (data.get("device_token") or "").strip()
    if not key:
        return jsonify({"ok": False, "error": "Введите лицензионный ключ"})
    if not device_token:
        device_token = make_device_token()

    licenses = load_licenses()
    if key not in licenses:
        return jsonify({"ok": False, "error": "Ключ не найден"})

    bound = licenses[key].get("device_token")
    if bound and bound != device_token:
        return jsonify({"ok": False, "error": "Ключ уже привязан к другому устройству"})

    # Привязываем устройство если ещё не привязан
    if not bound:
        licenses[key]["device_token"] = device_token
        save_licenses(licenses)

    # Создаём сессию
    token    = make_token()
    sessions = load_web_sessions()
    sessions[token] = {
        "key":          key,
        "device_token": device_token,
        "created_at":   datetime.now().isoformat(),
    }
    save_web_sessions(sessions)
    return jsonify({"ok": True, "token": token, "device_token": device_token})


@app.route("/api/auth/verify")
def api_auth_verify():
    token = request.headers.get("X-Token", "") or request.args.get("token", "")
    sessions = load_web_sessions()
    if token in sessions:
        sess = sessions[token]
        username = sess.get("username")
        license_key = None
        if username:
            accs = load_user_accounts()
            license_key = accs.get(username, {}).get("license_key")
        return jsonify({"ok": True, "username": username, "key": sess.get("key"), "license_key": license_key})
    return jsonify({"ok": False, "error": "unauthorized"}), 403


@app.route("/api/auth/logout", methods=["POST"])
def api_auth_logout():
    token    = request.headers.get("X-Token", "") or (request.get_json(force=True, silent=True) or {}).get("token", "")
    sessions = load_web_sessions()
    sessions.pop(token, None)
    save_web_sessions(sessions)
    return jsonify({"ok": True})


@app.route("/api/admin/licenses", methods=["GET"])
def api_admin_licenses_list():
    if request.headers.get("X-Admin") != ADMIN_MASTER:
        return jsonify({"ok": False, "error": "Неверный мастер-пароль"}), 403
    licenses = load_licenses()
    now = datetime.now()
    result = []
    for k, v in licenses.items():
        exp = v.get("expires_at")
        expired = False
        if exp:
            try:
                expired = datetime.fromisoformat(exp) < now
            except Exception:
                pass
        result.append({
            "key":          k,
            "note":         v.get("note", ""),
            "created_at":   v.get("created_at", ""),
            "device_token": v.get("device_token"),
            "bound":        bool(v.get("device_token")),
            "expires_at":   exp,
            "expired":      expired,
        })
    return jsonify({"ok": True, "licenses": result})


@app.route("/api/admin/licenses/add", methods=["POST"])
def api_admin_licenses_add():
    if request.headers.get("X-Admin") != ADMIN_MASTER:
        return jsonify({"ok": False, "error": "Неверный мастер-пароль"}), 403
    data     = request.get_json(force=True, silent=True) or {}
    key      = (data.get("key") or "").strip()
    note     = (data.get("note") or "").strip()
    if not key:
        return jsonify({"ok": False, "error": "Укажите ключ"})
    licenses = load_licenses()
    if key in licenses:
        return jsonify({"ok": False, "error": "Ключ уже существует"})
    days  = int(data.get("days")  or 0)
    hours = int(data.get("hours") or 0)
    expires_at = None
    if days or hours:
        expires_at = (datetime.now() + timedelta(days=days, hours=hours)).strftime("%Y-%m-%d %H:%M:%S")
    licenses[key] = {
        "note": note,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "expires_at": expires_at,
        "device_token": None,
    }
    save_licenses(licenses)
    return jsonify({"ok": True})


@app.route("/api/admin/licenses/delete", methods=["POST"])
def api_admin_licenses_delete():
    if request.headers.get("X-Admin") != ADMIN_MASTER:
        return jsonify({"ok": False, "error": "Неверный мастер-пароль"}), 403
    data = request.get_json(force=True, silent=True) or {}
    key  = (data.get("key") or "").strip()
    licenses = load_licenses()
    if key not in licenses:
        return jsonify({"ok": False, "error": "Ключ не найден"})
    licenses.pop(key)
    save_licenses(licenses)
    return jsonify({"ok": True})


@app.route("/api/admin/licenses/unbind", methods=["POST"])
def api_admin_licenses_unbind():
    """Сбросить привязку к устройству (для переноса на другой браузер)."""
    if request.headers.get("X-Admin") != ADMIN_MASTER:
        return jsonify({"ok": False, "error": "Неверный мастер-пароль"}), 403
    data = request.get_json(force=True, silent=True) or {}
    key  = (data.get("key") or "").strip()
    licenses = load_licenses()
    if key not in licenses:
        return jsonify({"ok": False, "error": "Ключ не найден"})
    licenses[key]["device_token"] = None
    save_licenses(licenses)
    return jsonify({"ok": True})


# ── Аккаунты пользователей ────────────────────────────────────────────────

@app.route("/api/account/register", methods=["POST"])
def api_account_register():
    import re as _re
    _clean_expired()
    data     = request.get_json(force=True, silent=True) or {}
    username = (data.get("username") or "").strip().lower()
    password = (data.get("password") or "").strip()
    email    = (data.get("email") or "").strip().lower()
    if not username or not password or not email:
        return jsonify({"ok": False, "error": "Укажите логин, email и пароль"})
    if len(username) < 3:
        return jsonify({"ok": False, "error": "Логин минимум 3 символа"})
    if len(password) < 4:
        return jsonify({"ok": False, "error": "Пароль минимум 4 символа"})
    if not _re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
        return jsonify({"ok": False, "error": "Неверный формат email"})
    accounts = load_user_accounts()
    if username in accounts:
        return jsonify({"ok": False, "error": "Пользователь уже существует"})
    for udata in accounts.values():
        if udata.get("email") == email:
            return jsonify({"ok": False, "error": "Email уже используется"})
    # Сохраняем pending-запись и отправляем код
    code = _make_code()
    _pending_regs[email] = {
        "username": username,
        "password_hash": hash_password(password),
        "code": code,
        "expires": time.time() + 900,  # 15 минут
        "attempts": 0,
    }
    cfg = _smtp_cfg()
    if not cfg["smtp_host"]:
        # SMTP не настроен — сразу создаём аккаунт (режим без подтверждения)
        accounts[username] = {
            "password_hash": hash_password(password),
            "email": email,
            "email_confirmed": False,
            "license_key": None,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
        save_user_accounts(accounts)
        _pending_regs.pop(email, None)
        token = make_token()
        sessions = load_web_sessions()
        sessions[token] = {"username": username, "created_at": datetime.now().isoformat()}
        save_web_sessions(sessions)
        return jsonify({"ok": True, "token": token, "username": username, "license_key": None, "confirmed": True})
    body = email_template("Подтверждение регистрации", f"""
      <p style="margin:0 0 14px;">Здравствуйте, <b style="color:#4ade80;">{username}</b>!</p>
      <p style="margin:0 0 18px;color:#9ca3af;">Ваш код активации:</p>
      <table width="100%" cellpadding="0" cellspacing="0" border="0">
        <tr><td align="center" bgcolor="#0c0e12" style="border-radius:10px;border:1px solid #1e3a2a;padding:16px 24px;">
          <span style="font-family:monospace;font-size:34px;font-weight:700;letter-spacing:10px;color:#4ade80;">{code}</span>
        </td></tr>
      </table>
      <p style="margin:18px 0 0;font-size:12px;color:#4b5563;">Код действителен 15 минут. Если вы не регистрировались — проигнорируйте это письмо.</p>
    """)
    ok, err = send_email(email, "Код активации TG Channel Creator", body)
    if not ok:
        _pending_regs.pop(email, None)
        return jsonify({"ok": False, "error": f"Не удалось отправить письмо: {err}"})
    log("info", f"Код активации отправлен на {email}")
    return jsonify({"ok": True, "pending": True, "email": email})


@app.route("/api/account/confirm", methods=["POST"])
def api_account_confirm():
    _clean_expired()
    data  = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    code  = (data.get("code") or "").strip()
    if not email or not code:
        return jsonify({"ok": False, "error": "Укажите email и код"})
    pending = _pending_regs.get(email)
    if not pending:
        return jsonify({"ok": False, "error": "Код не найден или истёк — запросите новый"})
    pending["attempts"] += 1
    if pending["attempts"] > 5:
        _pending_regs.pop(email, None)
        return jsonify({"ok": False, "error": "Слишком много попыток. Пройдите регистрацию заново"})
    if pending["code"] != code:
        left = 5 - pending["attempts"]
        return jsonify({"ok": False, "error": f"Неверный код (осталось попыток: {left})"})
    # Создаём аккаунт
    username = pending["username"]
    accounts = load_user_accounts()
    if username in accounts:
        _pending_regs.pop(email, None)
        return jsonify({"ok": False, "error": "Пользователь уже существует"})
    accounts[username] = {
        "password_hash": pending["password_hash"],
        "email": email,
        "email_confirmed": True,
        "license_key": None,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    save_user_accounts(accounts)
    _pending_regs.pop(email, None)
    token = make_token()
    sessions = load_web_sessions()
    sessions[token] = {"username": username, "created_at": datetime.now().isoformat()}
    save_web_sessions(sessions)
    log("success", f"Аккаунт {username} подтверждён по email")
    return jsonify({"ok": True, "token": token, "username": username, "license_key": None})


@app.route("/api/account/resend-code", methods=["POST"])
def api_account_resend_code():
    _clean_expired()
    data  = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    if not email:
        return jsonify({"ok": False, "error": "Укажите email"})
    pending = _pending_regs.get(email)
    if not pending:
        return jsonify({"ok": False, "error": "Сессия истекла — пройдите регистрацию заново"})
    code = _make_code()
    pending["code"] = code
    pending["expires"] = time.time() + 900
    pending["attempts"] = 0
    body = f"""
    <div style="font-family:sans-serif;max-width:480px;margin:0 auto;padding:30px 20px;background:#0c0e12;color:#e8eaf0;border-radius:12px;">
      <div style="font-size:22px;font-weight:700;color:#4ade80;margin-bottom:8px;">TG Channel Creator</div>
      <div style="font-size:15px;margin-bottom:24px;color:#8b90a0;">Подтверждение регистрации (повторная отправка)</div>
      <p style="margin-bottom:20px;">Новый код активации:</p>
      <div style="font-size:36px;font-weight:700;letter-spacing:10px;color:#4ade80;background:#13161c;border:1px solid rgba(74,222,128,.25);border-radius:10px;padding:18px 24px;text-align:center;font-family:monospace;">{code}</div>
      <p style="margin-top:20px;font-size:12px;color:#555a6a;">Код действителен 15 минут.</p>
    </div>
    """
    ok, err = send_email(email, "Новый код активации TG Channel Creator", body)
    if not ok:
        return jsonify({"ok": False, "error": f"Не удалось отправить письмо: {err}"})
    return jsonify({"ok": True})


@app.route("/api/account/forgot-password", methods=["POST"])
def api_account_forgot_password():
    import re as _re
    _clean_expired()
    data  = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    if not email or not _re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
        return jsonify({"ok": False, "error": "Укажите корректный email"})
    accounts = load_user_accounts()
    username = next((u for u, d in accounts.items() if d.get("email") == email), None)
    if not username:
        # Не раскрываем, существует ли email
        return jsonify({"ok": True})
    code = _make_code()
    _reset_tokens[email] = {
        "username": username,
        "code": code,
        "expires": time.time() + 900,
        "attempts": 0,
    }
    body = f"""
    <div style="font-family:sans-serif;max-width:480px;margin:0 auto;padding:30px 20px;background:#0c0e12;color:#e8eaf0;border-radius:12px;">
      <div style="font-size:22px;font-weight:700;color:#4ade80;margin-bottom:8px;">TG Channel Creator</div>
      <div style="font-size:15px;margin-bottom:24px;color:#8b90a0;">Восстановление пароля</div>
      <p style="margin-bottom:16px;">Здравствуйте, <b>{username}</b>!</p>
      <p style="margin-bottom:20px;">Код для сброса пароля:</p>
      <div style="font-size:36px;font-weight:700;letter-spacing:10px;color:#f59e0b;background:#13161c;border:1px solid rgba(245,158,11,.25);border-radius:10px;padding:18px 24px;text-align:center;font-family:monospace;">{code}</div>
      <p style="margin-top:20px;font-size:12px;color:#555a6a;">Код действителен 15 минут. Если вы не запрашивали сброс — проигнорируйте это письмо.</p>
    </div>
    """
    ok, err = send_email(email, "Сброс пароля TG Channel Creator", body)
    if not ok:
        _reset_tokens.pop(email, None)
        return jsonify({"ok": False, "error": f"Не удалось отправить письмо: {err}"})
    log("info", f"Код сброса пароля отправлен на {email}")
    return jsonify({"ok": True})


@app.route("/api/account/reset-password", methods=["POST"])
def api_account_reset_password():
    _clean_expired()
    data     = request.get_json(force=True, silent=True) or {}
    email    = (data.get("email") or "").strip().lower()
    code     = (data.get("code") or "").strip()
    new_pass = (data.get("password") or "").strip()
    if not email or not code or not new_pass:
        return jsonify({"ok": False, "error": "Заполните все поля"})
    if len(new_pass) < 4:
        return jsonify({"ok": False, "error": "Пароль минимум 4 символа"})
    token_data = _reset_tokens.get(email)
    if not token_data:
        return jsonify({"ok": False, "error": "Код не найден или истёк — запросите новый"})
    token_data["attempts"] += 1
    if token_data["attempts"] > 5:
        _reset_tokens.pop(email, None)
        return jsonify({"ok": False, "error": "Слишком много попыток. Запросите новый код"})
    if token_data["code"] != code:
        left = 5 - token_data["attempts"]
        return jsonify({"ok": False, "error": f"Неверный код (осталось попыток: {left})"})
    username = token_data["username"]
    accounts = load_user_accounts()
    if username not in accounts:
        return jsonify({"ok": False, "error": "Пользователь не найден"})
    accounts[username]["password_hash"] = hash_password(new_pass)
    save_user_accounts(accounts)
    _reset_tokens.pop(email, None)
    _login_attempts.pop(username, None)  # снимаем блокировку если была
    log("success", f"Пароль сброшен для {username}")
    return jsonify({"ok": True})


@app.route("/api/account/login", methods=["POST"])
def api_account_login():
    data      = request.get_json(force=True, silent=True) or {}
    login_str = (data.get("username") or "").strip().lower()
    password  = (data.get("password") or "").strip()
    if not login_str or not password:
        return jsonify({"ok": False, "error": "Укажите логин и пароль"})

    # Rate limiting: 5 попыток, блокировка 15 минут
    now = time.time()
    att = _login_attempts.get(login_str, {"count": 0, "blocked_until": 0.0})
    if att["blocked_until"] > now:
        remaining = int(att["blocked_until"] - now)
        mins, secs = divmod(remaining, 60)
        return jsonify({"ok": False, "error": f"Слишком много попыток. Подождите {mins}м {secs}с"})

    accounts = load_user_accounts()

    # Поиск по логину или email
    username = None
    if login_str in accounts:
        username = login_str
    elif "@" in login_str:
        for uname, udata in accounts.items():
            if udata.get("email", "").lower() == login_str:
                username = uname
                break

    if username is None or accounts[username]["password_hash"] != hash_password(password):
        att["count"] += 1
        if att["count"] >= 5:
            att["blocked_until"] = now + 900  # 15 минут
        _login_attempts[login_str] = att
        left = max(0, 5 - att["count"])
        msg = "Неверный логин или пароль"
        if left > 0:
            msg += f" (осталось попыток: {left})"
        else:
            msg = "Аккаунт заблокирован на 15 минут"
        return jsonify({"ok": False, "error": msg})

    if accounts[username].get("blocked"):
        return jsonify({"ok": False, "error": "Аккаунт заблокирован администратором"})

    _login_attempts.pop(login_str, None)  # сброс при успехе
    token    = make_token()
    sessions = load_web_sessions()
    sessions[token] = {"username": username, "created_at": datetime.now().isoformat()}
    save_web_sessions(sessions)
    lic_key = accounts[username].get("license_key")
    return jsonify({
        "ok": True, "token": token, "username": username,
        "created_at": accounts[username].get("created_at"),
        "license_key": lic_key,
        "expires_at": _key_expires_at(lic_key) if lic_key else None,
    })


def _key_expires_at(key: str) -> str | None:
    """Возвращает expires_at из licenses.json для данного ключа."""
    return load_licenses().get(key, {}).get("expires_at")


@app.route("/api/account/me")
def api_account_me():
    token    = request.headers.get("X-Token", "")
    sessions = load_web_sessions()
    if token not in sessions:
        return jsonify({"ok": False, "error": "unauthorized"}), 403
    username = sessions[token].get("username")
    accounts = load_user_accounts()
    if not username or username not in accounts:
        return jsonify({"ok": False, "error": "Пользователь не найден"}), 404
    lic_key = accounts[username].get("license_key")
    return jsonify({
        "ok": True,
        "username": username,
        "created_at": accounts[username].get("created_at"),
        "license_key": lic_key,
        "expires_at": _key_expires_at(lic_key) if lic_key else None,
    })


@app.route("/api/account/bind-key", methods=["POST"])
def api_account_bind_key():
    token    = request.headers.get("X-Token", "")
    sessions = load_web_sessions()
    if token not in sessions:
        return jsonify({"ok": False, "error": "Войдите в аккаунт"}), 403
    username = sessions[token].get("username")
    if not username:
        return jsonify({"ok": False, "error": "Сессия устарела, войдите снова"}), 403
    data     = request.get_json(force=True, silent=True) or {}
    key      = (data.get("key") or "").strip()
    if not key:
        return jsonify({"ok": False, "error": "Укажите ключ"})
    licenses = load_licenses()
    if key not in licenses:
        return jsonify({"ok": False, "error": "Ключ не найден"})
    accounts = load_user_accounts()
    # Проверяем, не занят ли ключ другим аккаунтом
    for uname, udata in accounts.items():
        if udata.get("license_key") == key and uname != username:
            return jsonify({"ok": False, "error": "Ключ уже привязан к другому аккаунту"})
    # Проверяем срок действия
    exp = licenses[key].get("expires_at")
    if exp:
        try:
            if datetime.fromisoformat(exp) < datetime.now():
                return jsonify({"ok": False, "error": "Срок действия ключа истёк"})
        except Exception:
            pass
    accounts[username]["license_key"] = key
    save_user_accounts(accounts)
    return jsonify({"ok": True, "license_key": key, "expires_at": exp})


@app.route("/api/account/change-password", methods=["POST"])
def api_account_change_password():
    token    = request.headers.get("X-Token", "")
    sessions = load_web_sessions()
    if token not in sessions:
        return jsonify({"ok": False, "error": "unauthorized"}), 403
    username = sessions[token].get("username")
    data     = request.get_json(force=True, silent=True) or {}
    old_pw   = (data.get("old_password") or "").strip()
    new_pw   = (data.get("new_password") or "").strip()
    if not old_pw or not new_pw:
        return jsonify({"ok": False, "error": "Укажите старый и новый пароль"})
    if len(new_pw) < 4:
        return jsonify({"ok": False, "error": "Новый пароль минимум 4 символа"})
    accounts = load_user_accounts()
    if not username or username not in accounts:
        return jsonify({"ok": False, "error": "Пользователь не найден"}), 404
    if accounts[username]["password_hash"] != hash_password(old_pw):
        return jsonify({"ok": False, "error": "Неверный текущий пароль"})
    accounts[username]["password_hash"] = hash_password(new_pw)
    save_user_accounts(accounts)
    return jsonify({"ok": True})


@app.route("/api/account/unbind-key", methods=["POST"])
def api_account_unbind_key():
    token    = request.headers.get("X-Token", "")
    sessions = load_web_sessions()
    if token not in sessions:
        return jsonify({"ok": False, "error": "unauthorized"}), 403
    username = sessions[token].get("username")
    accounts = load_user_accounts()
    if not username or username not in accounts:
        return jsonify({"ok": False, "error": "Пользователь не найден"}), 404
    accounts[username]["license_key"] = None
    save_user_accounts(accounts)
    return jsonify({"ok": True})


@app.route("/api/admin/accounts", methods=["GET"])
def api_admin_accounts():
    if request.headers.get("X-Admin") != ADMIN_MASTER:
        return jsonify({"ok": False, "error": "Неверный мастер-пароль"}), 403
    accounts = load_user_accounts()
    result = [
        {"username": u, "created_at": d.get("created_at"), "license_key": d.get("license_key")}
        for u, d in accounts.items()
    ]
    return jsonify({"ok": True, "accounts": result, "count": len(result)})


# ── Admin API ──────────────────────────────────────────────────────────────

def _admin_check():
    return request.headers.get("X-Admin") == ADMIN_MASTER

@app.route("/api/admin/auth", methods=["POST"])
def api_admin_auth():
    data = request.get_json(force=True, silent=True) or {}
    if data.get("password") == ADMIN_MASTER:
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "Неверный пароль"}), 403

@app.route("/api/admin/stats", methods=["GET"])
def api_admin_stats():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    accounts = load_user_accounts()
    licenses = load_licenses()
    sessions = load_web_sessions()
    now = datetime.now()
    today = now.date()
    week_ago = now - timedelta(days=7)
    active_keys = sum(
        1 for v in licenses.values()
        if not v.get("expires_at") or datetime.fromisoformat(v["expires_at"]) > now
    )
    free_keys = sum(1 for v in licenses.values() if not v.get("device_token"))
    # online = unique usernames with active sessions
    online_usernames = set(s.get("username") for s in sessions.values() if s.get("username"))
    reg_today = sum(
        1 for d in accounts.values()
        if d.get("created_at") and d["created_at"][:10] == str(today)
    )
    reg_week = sum(
        1 for d in accounts.values()
        if d.get("created_at") and datetime.fromisoformat(d["created_at"][:16]) >= week_ago
    )
    return jsonify({"ok": True, "stats": {
        "total_users": len(accounts),
        "verified": sum(1 for d in accounts.values() if d.get("email_confirmed")),
        "blocked": sum(1 for d in accounts.values() if d.get("blocked")),
        "online": len(online_usernames),
        "active_keys": active_keys,
        "free_keys": free_keys,
        "reg_today": reg_today,
        "reg_week": reg_week,
    }})

@app.route("/api/admin/online", methods=["GET"])
def api_admin_online():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    accounts = load_user_accounts()
    sessions = load_web_sessions()
    # group sessions by username
    by_user = {}
    for token, sdata in sessions.items():
        uname = sdata.get("username")
        if not uname:
            continue
        by_user.setdefault(uname, []).append({
            "token": token[:8] + "…",
            "full_token": token,
            "created_at": sdata.get("created_at"),
            "ip": sdata.get("ip", ""),
            "user_agent": sdata.get("user_agent", ""),
        })
    result = []
    for uname, sess_list in by_user.items():
        udata = accounts.get(uname, {})
        result.append({
            "username": uname,
            "email": udata.get("email", ""),
            "blocked": bool(udata.get("blocked")),
            "sessions": sess_list,
        })
    return jsonify({"ok": True, "online": result})

@app.route("/api/admin/users", methods=["GET"])
def api_admin_users():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    accounts = load_user_accounts()
    sessions = load_web_sessions()
    licenses = load_licenses()
    # count sessions per user
    sess_count = {}
    for sdata in sessions.values():
        uname = sdata.get("username")
        if uname:
            sess_count[uname] = sess_count.get(uname, 0) + 1
    now = datetime.now()
    result = []
    for uname, udata in accounts.items():
        lic_key = udata.get("license_key")
        lic_exp = None
        if lic_key and lic_key in licenses:
            lic_exp = licenses[lic_key].get("expires_at")
        result.append({
            "username": uname,
            "email": udata.get("email", ""),
            "email_verified": bool(udata.get("email_confirmed")),
            "blocked": bool(udata.get("blocked")),
            "license_key": lic_key,
            "license_expires": lic_exp,
            "created_at": udata.get("created_at", ""),
            "sessions": sess_count.get(uname, 0),
        })
    return jsonify({"ok": True, "users": result})

@app.route("/api/admin/users/reset-password", methods=["POST"])
def api_admin_reset_password():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    data = request.get_json(force=True, silent=True) or {}
    username = (data.get("username") or "").strip().lower()
    new_pw   = (data.get("new_password") or "").strip()
    if not username or not new_pw:
        return jsonify({"ok": False, "error": "Укажите username и новый пароль"})
    accounts = load_user_accounts()
    if username not in accounts:
        return jsonify({"ok": False, "error": "Пользователь не найден"})
    accounts[username]["password_hash"] = hash_password(new_pw)
    save_user_accounts(accounts)
    return jsonify({"ok": True})

@app.route("/api/admin/users/set-key", methods=["POST"])
def api_admin_set_key():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    data = request.get_json(force=True, silent=True) or {}
    username = (data.get("username") or "").strip().lower()
    key      = (data.get("key") or "").strip()
    accounts = load_user_accounts()
    if username not in accounts:
        return jsonify({"ok": False, "error": "Пользователь не найден"})
    accounts[username]["license_key"] = key or None
    save_user_accounts(accounts)
    return jsonify({"ok": True})

@app.route("/api/admin/users/block", methods=["POST"])
def api_admin_block_user():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    data = request.get_json(force=True, silent=True) or {}
    username = (data.get("username") or "").strip().lower()
    block    = bool(data.get("block", True))
    accounts = load_user_accounts()
    if username not in accounts:
        return jsonify({"ok": False, "error": "Пользователь не найден"})
    accounts[username]["blocked"] = block
    save_user_accounts(accounts)
    return jsonify({"ok": True})

@app.route("/api/admin/users/delete", methods=["POST"])
def api_admin_delete_user():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    data = request.get_json(force=True, silent=True) or {}
    username = (data.get("username") or "").strip().lower()
    accounts = load_user_accounts()
    if username not in accounts:
        return jsonify({"ok": False, "error": "Пользователь не найден"})
    accounts.pop(username)
    save_user_accounts(accounts)
    # revoke sessions
    sessions = load_web_sessions()
    to_del = [t for t, s in sessions.items() if s.get("username") == username]
    for t in to_del:
        sessions.pop(t)
    save_web_sessions(sessions)
    return jsonify({"ok": True})

@app.route("/api/admin/sessions", methods=["GET"])
def api_admin_sessions():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    sessions = load_web_sessions()
    result = [
        {
            "token": t[:8] + "…",
            "full_token": t,
            "username": s.get("username", ""),
            "created_at": s.get("created_at", ""),
            "ip": s.get("ip", ""),
            "user_agent": s.get("user_agent", ""),
        }
        for t, s in sessions.items()
    ]
    return jsonify({"ok": True, "sessions": result})

@app.route("/api/admin/sessions/revoke", methods=["POST"])
def api_admin_revoke_session():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    data = request.get_json(force=True, silent=True) or {}
    token = (data.get("token") or "").strip()
    sessions = load_web_sessions()
    if token not in sessions:
        return jsonify({"ok": False, "error": "Сессия не найдена"})
    sessions.pop(token)
    save_web_sessions(sessions)
    return jsonify({"ok": True})

@app.route("/api/admin/sessions/revoke-user", methods=["POST"])
def api_admin_revoke_user_sessions():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    data = request.get_json(force=True, silent=True) or {}
    username = (data.get("username") or "").strip().lower()
    sessions = load_web_sessions()
    to_del = [t for t, s in sessions.items() if s.get("username") == username]
    for t in to_del:
        sessions.pop(t)
    save_web_sessions(sessions)
    return jsonify({"ok": True, "revoked": len(to_del)})

@app.route("/api/admin/licenses/extend", methods=["POST"])
def api_admin_licenses_extend():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    data = request.get_json(force=True, silent=True) or {}
    key  = (data.get("key") or "").strip()
    days = int(data.get("days") or 0)
    licenses = load_licenses()
    if key not in licenses:
        return jsonify({"ok": False, "error": "Ключ не найден"})
    exp = licenses[key].get("expires_at")
    base = datetime.fromisoformat(exp) if exp else datetime.now()
    if base < datetime.now():
        base = datetime.now()
    licenses[key]["expires_at"] = (base + timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
    save_licenses(licenses)
    return jsonify({"ok": True})

@app.route("/api/admin/notify", methods=["POST"])
def api_admin_notify():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    data    = request.get_json(force=True, silent=True) or {}
    target  = (data.get("target") or "").strip().lower()
    subject = (data.get("subject") or "Сообщение от администратора").strip()
    body    = (data.get("body") or "").strip()
    html = email_template("Сообщение от администратора", f"""
      <div style="font-size:14px;line-height:1.8;color:#d1d5db;">{body.replace(chr(10), "<br>")}</div>
    """)
    accounts = load_user_accounts()

    # Рассылка всем
    if target == "all":
        sent, failed = 0, 0
        for udata in accounts.values():
            email = udata.get("email")
            if not email:
                continue
            ok, _ = send_email(email, subject, html)
            if ok:
                sent += 1
            else:
                failed += 1
        return jsonify({"ok": True, "sent": sent, "failed": failed})

    # Один пользователь или прямой email
    email = None
    if target in accounts:
        email = accounts[target].get("email")
    elif "@" in target:
        email = target
    if not email:
        return jsonify({"ok": False, "error": "Email пользователя не найден"})
    ok, err = send_email(email, subject, html)
    if not ok:
        return jsonify({"ok": False, "error": err})
    return jsonify({"ok": True, "sent": 1, "failed": 0})

@app.route("/api/admin/settings/smtp", methods=["GET", "POST"])
def api_admin_settings_smtp():
    if not _admin_check():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    if request.method == "GET":
        return jsonify({"ok": True,
            "smtp_host": settings.get("smtp_host", ""),
            "smtp_port": settings.get("smtp_port", 587),
            "smtp_user": settings.get("smtp_user", ""),
            "smtp_password": settings.get("smtp_password", ""),
            "smtp_from": settings.get("smtp_from_name", "TG Channel Creator"),
            "smtp_tls": settings.get("smtp_tls", True),
        })
    data = request.get_json(force=True, silent=True) or {}
    settings["smtp_host"]      = data.get("smtp_host", settings["smtp_host"])
    settings["smtp_port"]      = int(data.get("smtp_port", settings["smtp_port"]))
    settings["smtp_user"]      = data.get("smtp_user", settings["smtp_user"])
    settings["smtp_from_name"] = data.get("smtp_from", settings["smtp_from_name"])
    settings["smtp_tls"]       = bool(data.get("smtp_tls", settings["smtp_tls"]))
    if data.get("smtp_password"):
        settings["smtp_password"] = data["smtp_password"]
    return jsonify({"ok": True})


# ── Статика ────────────────────────────────────────────────────────────────

@app.route("/admin")
def serve_admin():
    return send_from_directory(".", "admin.html")

@app.route("/")
def serve_index():
    return send_from_directory(".", "index.html")


if __name__ == "__main__":
    import socket as _socket
    from werkzeug.serving import make_server

    _port = int(os.getenv("PORT", 5000))

    print("=" * 52)
    print("  TG Channel Creator — Web API")
    print(f"  http://localhost:{_port}")
    print("=" * 52)
    _sock = _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM)
    _sock.setsockopt(_socket.SOL_SOCKET, _socket.SO_REUSEADDR, 1)
    try:
        _sock.setsockopt(_socket.SOL_SOCKET, _socket.SO_REUSEPORT, 1)
    except AttributeError:
        pass
    _sock.bind(("0.0.0.0", _port))
    _sock.listen(128)
    srv = make_server("0.0.0.0", _port, app, threaded=True, fd=_sock.fileno())
    srv.serve_forever()
